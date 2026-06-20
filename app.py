from __future__ import annotations

import csv
import io
import os
import shutil
import sys
import sqlite3
from datetime import datetime
from functools import wraps
from pathlib import Path

from flask import (
    Flask,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

try:
    from openpyxl import load_workbook
except Exception:  # Excel 导入是增强功能，没有安装 openpyxl 时仍可运行系统。
    load_workbook = None


def resource_base_dir() -> Path:
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS)
    return Path(__file__).resolve().parent


def writable_base_dir() -> Path:
    if not getattr(sys, "frozen", False):
        return Path(__file__).resolve().parent
    executable = Path(sys.executable).resolve()
    app_bundle = next((parent for parent in executable.parents if parent.suffix == ".app"), None)
    if app_bundle:
        return app_bundle.parent / "企业采购仓库销售管理系统数据"
    return executable.parent / "企业采购仓库销售管理系统数据"


BASE_DIR = resource_base_dir()
WRITE_DIR = writable_base_dir()
DATA_DIR = WRITE_DIR / "data"
UPLOAD_DIR = WRITE_DIR / "uploads"
BACKUP_DIR = WRITE_DIR / "backups"
DB_PATH = DATA_DIR / "app.db"

app = Flask(
    __name__,
    template_folder=str(BASE_DIR / "templates"),
    static_folder=str(BASE_DIR / "static"),
)
app.config["SECRET_KEY"] = os.environ.get("APP_SECRET_KEY", "change-this-local-secret")
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024

ALLOWED_UPLOADS = {"jpg", "jpeg", "png", "pdf", "doc", "docx", "xls", "xlsx"}
IMPORT_FIELDS = {
    "suppliers": ["name", "level", "contact", "phone", "address", "account_info", "company_type", "settlement", "remark"],
    "customers": ["name", "level", "contact", "phone", "address", "account_info", "company_type", "settlement", "total_sales", "received_amount", "remark"],
    "products": ["code", "name", "description", "unit", "current_stock", "safe_stock", "equipment", "part_type", "warranty_until", "origin", "purchase_price", "sale_price", "remark"],
    "purchases": ["order_no", "supplier_code", "supplier_name", "order_date", "expected_date", "product_code", "name", "description", "unit", "quantity", "unit_price", "equipment", "part_type", "warranty_until", "origin", "remark"],
    "sales": ["order_no", "order_date", "customer_code", "product_code", "quantity", "unit_price", "delivery_date", "remark"],
}

FIELD_ALIASES = {
    "code": ["编号", "供应商编号", "客户编号", "备件编号", "产品编号"],
    "name": ["名称", "供应商名称", "客户名称", "产品名称"],
    "level": ["等级"],
    "contact": ["联系人"],
    "phone": ["电话", "联系电话"],
    "address": ["地址", "公司地址"],
    "account_info": ["账户信息"],
    "company_type": ["公司类型"],
    "settlement": ["结算时间", "结算方式"],
    "total_sales": ["应付金额", "累计销售额"],
    "received_amount": ["实收金额"],
    "description": ["技术描述", "规格型号"],
    "unit": ["单位"],
    "current_stock": ["当前库存", "初始库存"],
    "safe_stock": ["安全库存"],
    "equipment": ["应用设备", "设备名称"],
    "part_type": ["备件类型", "产品类型"],
    "warranty_until": ["质保时间", "质保到期日"],
    "origin": ["产地"],
    "purchase_price": ["采购价", "采购单价"],
    "sale_price": ["销售价", "销售单价"],
    "order_no": ["采购单号", "销售单号", "单号"],
    "order_date": ["采购日期", "销售日期", "日期"],
    "supplier_code": ["供应商编号"],
    "supplier_name": ["供应商", "供应商名称"],
    "customer_code": ["客户编号"],
    "product_code": ["备件编号", "产品编号"],
    "quantity": ["数量"],
    "unit_price": ["单价"],
    "delivery_date": ["货期"],
    "expected_date": ["到货时间", "预计到货日期", "要求到货日期"],
    "remark": ["备注"],
}

EXCEL_TEMPLATE_BLOCKS = {
    "suppliers": {"sheet": "采购", "title": "供应商录入模版"},
    "purchases": {"sheet": "采购", "title": "采购订单版"},
    "products": {"sheet": "仓库", "title": "产品模版"},
    "customers": {"sheet": "销售", "title": "客户录入模版"},
    "sales": {"sheet": "销售", "title": "销售订单模版"},
}


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def today_code_date() -> str:
    return datetime.now().strftime("%Y%m%d")


def ensure_dirs() -> None:
    for folder in (DATA_DIR, UPLOAD_DIR, BACKUP_DIR):
        folder.mkdir(parents=True, exist_ok=True)


def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def row_to_dict(row: sqlite3.Row | None) -> dict | None:
    return dict(row) if row else None


def rows_to_list(rows) -> list[dict]:
    return [dict(row) for row in rows]


def execute_schema() -> None:
    ensure_dirs()
    with db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL CHECK(role IN ('admin','user')),
                display_name TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS suppliers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                level TEXT DEFAULT 'B',
                contact TEXT,
                phone TEXT,
                address TEXT,
                account_info TEXT,
                company_type TEXT,
                settlement TEXT,
                document_path TEXT,
                photo_path TEXT,
                remark TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS supplier_attachments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                supplier_id INTEGER NOT NULL,
                file_name TEXT NOT NULL,
                file_path TEXT NOT NULL,
                file_type TEXT,
                uploaded_at TEXT NOT NULL,
                FOREIGN KEY(supplier_id) REFERENCES suppliers(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS customers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                level TEXT DEFAULT 'B',
                contact TEXT,
                phone TEXT,
                address TEXT,
                account_info TEXT,
                company_type TEXT,
                settlement TEXT,
                document_path TEXT,
                photo_path TEXT,
                total_sales REAL DEFAULT 0,
                received_amount REAL DEFAULT 0,
                remark TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                description TEXT,
                unit TEXT,
                current_stock REAL DEFAULT 0,
                safe_stock REAL DEFAULT 0,
                photo_path TEXT,
                spec_path TEXT,
                equipment TEXT,
                part_type TEXT,
                warranty_until TEXT,
                origin TEXT,
                purchase_price REAL DEFAULT 0,
                sale_price REAL DEFAULT 0,
                remark TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS purchase_orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_no TEXT NOT NULL UNIQUE,
                supplier_id INTEGER NOT NULL,
                order_date TEXT NOT NULL,
                buyer TEXT NOT NULL,
                product_id INTEGER NOT NULL,
                quantity REAL NOT NULL,
                unit_price REAL NOT NULL,
                amount REAL NOT NULL,
                status TEXT NOT NULL DEFAULT 'approved',
                remark TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(supplier_id) REFERENCES suppliers(id),
                FOREIGN KEY(product_id) REFERENCES products(id)
            );

            CREATE TABLE IF NOT EXISTS sales_orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_no TEXT NOT NULL UNIQUE,
                customer_id INTEGER NOT NULL,
                order_date TEXT NOT NULL,
                seller TEXT NOT NULL,
                product_id INTEGER NOT NULL,
                quantity REAL NOT NULL,
                unit_price REAL NOT NULL,
                amount REAL NOT NULL,
                delivery_date TEXT,
                status TEXT NOT NULL DEFAULT 'approved',
                remark TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(customer_id) REFERENCES customers(id),
                FOREIGN KEY(product_id) REFERENCES products(id)
            );

            -- V2：订单采用“主表 + 多条明细”结构，支持分批到货和分批发货。
            CREATE TABLE IF NOT EXISTS purchase_order_headers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_no TEXT NOT NULL UNIQUE,
                supplier_id INTEGER NOT NULL,
                order_date TEXT NOT NULL,
                expected_date TEXT,
                buyer TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending',
                remark TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                completed_at TEXT,
                FOREIGN KEY(supplier_id) REFERENCES suppliers(id)
            );

            CREATE TABLE IF NOT EXISTS purchase_order_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER NOT NULL,
                product_id INTEGER NOT NULL,
                quantity REAL NOT NULL,
                unit_price REAL NOT NULL,
                amount REAL NOT NULL,
                received_quantity REAL NOT NULL DEFAULT 0,
                remark TEXT,
                FOREIGN KEY(order_id) REFERENCES purchase_order_headers(id) ON DELETE CASCADE,
                FOREIGN KEY(product_id) REFERENCES products(id)
            );

            CREATE TABLE IF NOT EXISTS sales_order_headers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_no TEXT NOT NULL UNIQUE,
                customer_id INTEGER NOT NULL,
                order_date TEXT NOT NULL,
                delivery_date TEXT,
                seller TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending',
                remark TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                completed_at TEXT,
                FOREIGN KEY(customer_id) REFERENCES customers(id)
            );

            CREATE TABLE IF NOT EXISTS sales_order_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER NOT NULL,
                product_id INTEGER NOT NULL,
                quantity REAL NOT NULL,
                unit_price REAL NOT NULL,
                amount REAL NOT NULL,
                shipped_quantity REAL NOT NULL DEFAULT 0,
                remark TEXT,
                FOREIGN KEY(order_id) REFERENCES sales_order_headers(id) ON DELETE CASCADE,
                FOREIGN KEY(product_id) REFERENCES products(id)
            );

            CREATE TABLE IF NOT EXISTS stock_movements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER NOT NULL,
                movement_time TEXT NOT NULL,
                movement_type TEXT NOT NULL,
                quantity REAL NOT NULL,
                operator TEXT NOT NULL,
                source_no TEXT,
                remark TEXT,
                FOREIGN KEY(product_id) REFERENCES products(id)
            );

            CREATE TABLE IF NOT EXISTS system_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                log_time TEXT NOT NULL,
                username TEXT NOT NULL,
                action TEXT NOT NULL,
                detail TEXT,
                ip TEXT
            );
            """
        )
        sales_columns = {row["name"] for row in conn.execute("PRAGMA table_info(sales_orders)")}
        if "status" not in sales_columns:
            conn.execute("ALTER TABLE sales_orders ADD COLUMN status TEXT NOT NULL DEFAULT 'approved'")
        migrate_legacy_orders(conn)
        if conn.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
            conn.execute(
                "INSERT INTO users (username, password_hash, role, display_name, created_at) VALUES (?,?,?,?,?)",
                ("admin", generate_password_hash("admin123"), "admin", "系统管理员", now_text()),
            )
            conn.execute(
                "INSERT INTO users (username, password_hash, role, display_name, created_at) VALUES (?,?,?,?,?)",
                ("user", generate_password_hash("user123"), "user", "普通用户", now_text()),
            )


def migrate_legacy_orders(conn: sqlite3.Connection) -> None:
    """把旧版单行订单迁移为已完成历史整单；库存不重复变化。"""
    purchases = conn.execute("SELECT * FROM purchase_orders ORDER BY id").fetchall()
    for old in purchases:
        if conn.execute("SELECT 1 FROM purchase_order_headers WHERE order_no=?", (old["order_no"],)).fetchone():
            continue
        cursor = conn.execute(
            """
            INSERT INTO purchase_order_headers
            (order_no,supplier_id,order_date,expected_date,buyer,status,remark,created_at,updated_at,completed_at)
            VALUES (?,?,?,?,?,?,?,?,?,?)
            """,
            (old["order_no"], old["supplier_id"], old["order_date"], old["order_date"], old["buyer"],
             "completed", old["remark"], old["created_at"], old["updated_at"], old["updated_at"]),
        )
        conn.execute(
            """
            INSERT INTO purchase_order_items
            (order_id,product_id,quantity,unit_price,amount,received_quantity,remark)
            VALUES (?,?,?,?,?,?,?)
            """,
            (cursor.lastrowid, old["product_id"], old["quantity"], old["unit_price"],
             old["amount"], old["quantity"], "旧版采购记录迁移"),
        )

    sales = conn.execute("SELECT * FROM sales_orders ORDER BY id").fetchall()
    for old in sales:
        if conn.execute("SELECT 1 FROM sales_order_headers WHERE order_no=?", (old["order_no"],)).fetchone():
            continue
        cancelled = old["status"] == "voided"
        cursor = conn.execute(
            """
            INSERT INTO sales_order_headers
            (order_no,customer_id,order_date,delivery_date,seller,status,remark,created_at,updated_at,completed_at)
            VALUES (?,?,?,?,?,?,?,?,?,?)
            """,
            (old["order_no"], old["customer_id"], old["order_date"], old["delivery_date"], old["seller"],
             "cancelled" if cancelled else "completed", old["remark"], old["created_at"], old["updated_at"],
             old["updated_at"] if not cancelled else None),
        )
        conn.execute(
            """
            INSERT INTO sales_order_items
            (order_id,product_id,quantity,unit_price,amount,shipped_quantity,remark)
            VALUES (?,?,?,?,?,?,?)
            """,
            (cursor.lastrowid, old["product_id"], old["quantity"], old["unit_price"], old["amount"],
             0 if cancelled else old["quantity"], "旧版销售记录迁移"),
        )


def current_user() -> dict | None:
    return session.get("user")


def login_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not current_user():
            return jsonify({"ok": False, "message": "请先登录"}), 401
        return func(*args, **kwargs)

    return wrapper


def admin_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        user = current_user()
        if not user:
            return jsonify({"ok": False, "message": "请先登录"}), 401
        if user["role"] != "admin":
            return jsonify({"ok": False, "message": "当前账号没有管理员权限"}), 403
        return func(*args, **kwargs)

    return wrapper


def log_action(action: str, detail: str = "") -> None:
    user = current_user() or {"username": "system"}
    with db() as conn:
        conn.execute(
            "INSERT INTO system_logs (log_time, username, action, detail, ip) VALUES (?,?,?,?,?)",
            (now_text(), user["username"], action, detail, request.remote_addr if request else ""),
        )


def automatic_backup_once_per_day() -> None:
    """每天第一次访问系统时自动备份一次，避免引入复杂定时任务。"""
    if not DB_PATH.exists():
        return
    ensure_dirs()
    today = datetime.now().strftime("%Y%m%d")
    marker = BACKUP_DIR / f"auto_backup_{today}.db"
    if marker.exists():
        return
    shutil.copy2(DB_PATH, marker)
    with db() as conn:
        conn.execute(
            "INSERT INTO system_logs (log_time, username, action, detail, ip) VALUES (?,?,?,?,?)",
            (now_text(), "system", "自动备份", marker.name, request.remote_addr if request else ""),
        )


@app.before_request
def before_request_backup_check():
    if request.path.startswith("/static"):
        return
    automatic_backup_once_per_day()


def next_code(conn: sqlite3.Connection, table: str, prefix: str, column: str = "code") -> str:
    last = conn.execute(
        f"SELECT {column} FROM {table} WHERE {column} LIKE ? ORDER BY {column} DESC LIMIT 1",
        (f"{prefix}%",),
    ).fetchone()
    number = 1
    if last:
        try:
            number = int(str(last[column]).replace(prefix, "")) + 1
        except ValueError:
            number = 1
    return f"{prefix}{number:04d}"


def next_order_no(conn: sqlite3.Connection, table: str, prefix: str) -> str:
    date_part = today_code_date()
    full_prefix = f"{prefix}{date_part}"
    last = conn.execute(
        f"SELECT order_no FROM {table} WHERE order_no LIKE ? ORDER BY order_no DESC LIMIT 1",
        (f"{full_prefix}%",),
    ).fetchone()
    number = 1
    if last:
        number = int(last["order_no"].replace(full_prefix, "")) + 1
    return f"{full_prefix}{number:04d}"


def save_upload(field_name: str, subfolder: str) -> str | None:
    file = request.files.get(field_name)
    if not file or not file.filename:
        return None
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ALLOWED_UPLOADS:
        raise ValueError(f"不支持的附件类型：{ext}")
    folder = UPLOAD_DIR / subfolder
    folder.mkdir(parents=True, exist_ok=True)
    filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{secure_filename(file.filename)}"
    target = folder / filename
    file.save(target)
    return str(target.relative_to(WRITE_DIR))


def save_uploads(field_name: str, subfolder: str) -> list[dict]:
    saved_files = []
    for file in request.files.getlist(field_name):
        if not file or not file.filename:
            continue
        ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
        if ext not in ALLOWED_UPLOADS:
            raise ValueError(f"不支持的附件类型：{ext}")
        folder = UPLOAD_DIR / subfolder
        folder.mkdir(parents=True, exist_ok=True)
        safe_name = secure_filename(file.filename)
        filename = f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{safe_name}"
        target = folder / filename
        file.save(target)
        saved_files.append({
            "file_name": file.filename,
            "file_path": str(target.relative_to(WRITE_DIR)),
            "file_type": ext,
        })
    return saved_files


def delete_uploaded_file(relative_path: str | None) -> None:
    """删除系统保存的上传文件；只允许删除当前项目数据目录里的文件。"""
    if not relative_path:
        return
    target = (WRITE_DIR / relative_path).resolve()
    if WRITE_DIR.resolve() in target.parents and target.exists():
        target.unlink()


def payload() -> dict:
    if request.content_type and request.content_type.startswith("multipart/form-data"):
        return request.form.to_dict()
    return request.get_json(silent=True) or {}


def number_value(data: dict, key: str, default: float = 0) -> float:
    value = data.get(key, default)
    if value in ("", None):
        return default
    return float(value)


def save_purchase_product(conn: sqlite3.Connection, data: dict, unit_price: float, photo_path: str | None = None) -> int:
    """按采购 Excel 字段保存产品资料，返回产品 ID。"""
    product_code = (data.get("product_code") or "").strip()
    product_name = (data.get("product_name") or data.get("name") or "").strip()
    product = conn.execute("SELECT * FROM products WHERE code=?", (product_code,)).fetchone()
    if product:
        conn.execute(
            """
            UPDATE products
            SET name=?,description=?,unit=?,photo_path=COALESCE(?, photo_path),equipment=?,part_type=?,warranty_until=?,origin=?,purchase_price=?,updated_at=?
            WHERE id=?
            """,
            (product_name, data.get("description"), data.get("unit"), photo_path, data.get("equipment"), data.get("part_type"), data.get("warranty_until"), data.get("origin"), unit_price, now_text(), product["id"]),
        )
        return product["id"]
    cursor = conn.execute(
        """
        INSERT INTO products (code,name,description,unit,current_stock,safe_stock,photo_path,equipment,part_type,warranty_until,origin,purchase_price,sale_price,remark,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (product_code, product_name, data.get("description"), data.get("unit"), 0, 0, photo_path, data.get("equipment"), data.get("part_type"), data.get("warranty_until"), data.get("origin"), unit_price, 0, "采购单自动建立", now_text(), now_text()),
    )
    return cursor.lastrowid


def find_purchase_supplier(conn: sqlite3.Connection, data: dict) -> sqlite3.Row:
    supplier_id = data.get("supplier_id")
    if supplier_id:
        supplier = conn.execute("SELECT id FROM suppliers WHERE id=?", (supplier_id,)).fetchone()
        if supplier:
            return supplier
    supplier_code = (data.get("supplier_code") or "").strip()
    if supplier_code:
        supplier = conn.execute("SELECT id FROM suppliers WHERE code=?", (supplier_code,)).fetchone()
        if supplier:
            return supplier
    supplier_name = (data.get("supplier_name") or "").strip()
    if supplier_name:
        supplier = conn.execute("SELECT id FROM suppliers WHERE name=?", (supplier_name,)).fetchone()
        if supplier:
            return supplier
    return get_or_create_default_supplier(conn)


def import_value(item: dict, field: str, default=""):
    if field in item and item[field] not in (None, ""):
        return item[field]
    for alias in FIELD_ALIASES.get(field, []):
        if alias in item and item[alias] not in (None, ""):
            return item[alias]
    return default


def get_or_create_default_supplier(conn: sqlite3.Connection) -> sqlite3.Row:
    supplier = conn.execute("SELECT id FROM suppliers WHERE code=?", ("SUP-DEFAULT",)).fetchone()
    if supplier:
        return supplier
    conn.execute(
        """
        INSERT INTO suppliers (code,name,level,contact,phone,address,account_info,company_type,settlement,remark,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        ("SUP-DEFAULT", "未指定供应商", "C", "", "", "", "", "其它", "", "采购订单导入自动建立", now_text(), now_text()),
    )
    return conn.execute("SELECT id FROM suppliers WHERE code=?", ("SUP-DEFAULT",)).fetchone()


def get_or_create_default_customer(conn: sqlite3.Connection) -> sqlite3.Row:
    customer = conn.execute("SELECT id FROM customers WHERE code=?", ("CUS-DEFAULT",)).fetchone()
    if customer:
        return customer
    conn.execute(
        """
        INSERT INTO customers (code,name,level,contact,phone,address,account_info,company_type,settlement,total_sales,received_amount,remark,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        ("CUS-DEFAULT", "未指定客户", "C", "", "", "", "", "其它", "", 0, 0, "销售订单导入自动建立", now_text(), now_text()),
    )
    return conn.execute("SELECT id FROM customers WHERE code=?", ("CUS-DEFAULT",)).fetchone()


def read_excel_template_rows(file, table: str) -> list[dict]:
    if load_workbook is None:
        raise RuntimeError("未安装 openpyxl，不能导入 Excel，请先安装 requirements.txt")
    workbook = load_workbook(file, data_only=True)
    block = EXCEL_TEMPLATE_BLOCKS.get(table)
    sheet = workbook[block["sheet"]] if block and block["sheet"] in workbook.sheetnames else workbook.active

    title_row = None
    if block:
        for row in sheet.iter_rows():
            values = [cell.value for cell in row]
            if block["title"] in values:
                title_row = row[0].row
                break

    # 单独导入模板文件第 1 行是说明，第 2 行是表头；原总模板文件则是“标题下一行”为表头。
    header_row = (title_row + 1) if title_row else 1
    first_row_values = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    second_row_values = [sheet.cell(2, col).value for col in range(1, sheet.max_column + 1)]
    known_headers = {alias for aliases in FIELD_ALIASES.values() for alias in aliases}
    if not any(value in known_headers for value in first_row_values) and any(value in known_headers for value in second_row_values):
        header_row = 2
    headers = [sheet.cell(header_row, col).value for col in range(1, sheet.max_column + 1)]
    rows: list[dict] = []
    for row_index in range(header_row + 1, sheet.max_row + 1):
        item = {}
        for col_index, header in enumerate(headers, start=1):
            if header:
                item[str(header).strip()] = sheet.cell(row_index, col_index).value
        if any(value not in (None, "") for value in item.values()):
            rows.append(item)
    return rows


def date_range_from_request() -> tuple[str, str]:
    period = request.args.get("period") or "custom"
    today = datetime.now()
    if period == "day":
        return today.strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d")
    if period == "week":
        start = today.fromordinal(today.toordinal() - today.weekday())
        return start.strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d")
    if period == "month":
        return today.strftime("%Y-%m-01"), today.strftime("%Y-%m-%d")
    if period == "quarter":
        quarter_start_month = ((today.month - 1) // 3) * 3 + 1
        return today.replace(month=quarter_start_month, day=1).strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d")
    if period == "year":
        return today.strftime("%Y-01-01"), today.strftime("%Y-%m-%d")
    return request.args.get("start") or "1900-01-01", request.args.get("end") or "2999-12-31"


@app.route("/")
def index():
    if not current_user():
        return redirect(url_for("login_page"))
    return render_template("index.html", user=current_user())


@app.route("/login")
def login_page():
    return render_template("login.html")


@app.post("/api/login")
def api_login():
    data = payload()
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    with db() as conn:
        user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
    if not user or not check_password_hash(user["password_hash"], password):
        return jsonify({"ok": False, "message": "用户名或密码不正确"}), 400
    session["user"] = {
        "id": user["id"],
        "username": user["username"],
        "role": user["role"],
        "display_name": user["display_name"],
    }
    log_action("登录", f"{username} 登录系统")
    return jsonify({"ok": True, "user": session["user"]})


@app.post("/api/logout")
def api_logout():
    log_action("退出", "退出系统")
    session.clear()
    return jsonify({"ok": True})


@app.get("/api/session")
def api_session():
    return jsonify({"ok": True, "user": current_user()})


@app.get("/api/dashboard")
@login_required
def dashboard():
    month_start = datetime.now().strftime("%Y-%m-01")
    with db() as conn:
        product_total = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
        supplier_total = conn.execute("SELECT COUNT(*) FROM suppliers").fetchone()[0]
        customer_total = conn.execute("SELECT COUNT(*) FROM customers").fetchone()[0]
        stock_value = conn.execute("SELECT COALESCE(SUM(current_stock * purchase_price),0) FROM products").fetchone()[0]
        month_purchase = conn.execute(
            """SELECT COALESCE(SUM(poi.amount),0) FROM purchase_order_headers poh
               JOIN purchase_order_items poi ON poi.order_id=poh.id WHERE poh.order_date>=?""", (month_start,)
        ).fetchone()[0]
        month_sales = conn.execute(
            """SELECT COALESCE(SUM(soi.amount),0) FROM sales_order_headers soh
               JOIN sales_order_items soi ON soi.order_id=soh.id WHERE soh.order_date>=? AND soh.status!='cancelled'""", (month_start,)
        ).fetchone()[0]
        low_stock = conn.execute("SELECT COUNT(*) FROM products WHERE current_stock <= safe_stock").fetchone()[0]
        warranty_soon = conn.execute(
            "SELECT COUNT(*) FROM products WHERE warranty_until IS NOT NULL AND warranty_until!='' AND warranty_until <= date('now','+30 day')"
        ).fetchone()[0]
        recent_low = rows_to_list(conn.execute(
            "SELECT code, name, current_stock, safe_stock FROM products WHERE current_stock <= safe_stock ORDER BY current_stock ASC LIMIT 8"
        ))
        overdue_purchases = conn.execute(
            """SELECT COUNT(*) FROM purchase_order_headers WHERE status!='completed' AND expected_date!=''
               AND date(expected_date)<date('now','localtime')"""
        ).fetchone()[0]
        overdue_sales = conn.execute(
            """SELECT COUNT(*) FROM sales_order_headers WHERE status NOT IN ('completed','cancelled') AND delivery_date!=''
               AND date(delivery_date)<date('now','localtime')"""
        ).fetchone()[0]
    return jsonify({
        "ok": True,
        "cards": {
            "product_total": product_total,
            "supplier_total": supplier_total,
            "customer_total": customer_total,
            "stock_value": stock_value,
            "month_purchase": month_purchase,
            "month_sales": month_sales,
            "month_profit": month_sales - month_purchase,
            "low_stock": low_stock,
            "warranty_soon": warranty_soon,
            "overdue_purchases": overdue_purchases,
            "overdue_sales": overdue_sales,
        },
        "low_stock_items": recent_low,
    })


def list_table(table: str, search_columns: list[str], extra_select: str = "*", where_extra: str = "", order_by: str = "id DESC"):
    q = (request.args.get("q") or "").strip()
    params: list = []
    where = []
    if q:
        where.append("(" + " OR ".join([f"{col} LIKE ?" for col in search_columns]) + ")")
        params.extend([f"%{q}%"] * len(search_columns))
    if where_extra:
        where.append(where_extra)
    where_sql = "WHERE " + " AND ".join(where) if where else ""
    with db() as conn:
        rows = rows_to_list(conn.execute(f"SELECT {extra_select} FROM {table} {where_sql} ORDER BY {order_by}", params))
    return jsonify({"ok": True, "items": rows})


@app.get("/api/suppliers")
@login_required
def suppliers_list():
    return list_table(
        "suppliers s",
        ["s.code", "s.name", "s.contact"],
        "s.*, (SELECT COUNT(*) FROM supplier_attachments sa WHERE sa.supplier_id=s.id) AS attachment_count",
        order_by="s.id DESC",
    )


@app.post("/api/suppliers")
@admin_required
def suppliers_create():
    data = payload()
    if not (data.get("name") or "").strip():
        return jsonify({"ok": False, "message": "供应商名称必填"}), 400
    try:
        document_path = save_upload("document", "suppliers")
        photo_path = save_upload("photo", "suppliers")
        attachments = save_uploads("attachments", "suppliers")
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    with db() as conn:
        code = next_code(conn, "suppliers", "SUP")
        cursor = conn.execute(
            """
            INSERT INTO suppliers (code,name,level,contact,phone,address,account_info,company_type,settlement,document_path,photo_path,remark,created_at,updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (code, data.get("name"), data.get("level"), data.get("contact"), data.get("phone"), data.get("address"), data.get("account_info"), data.get("company_type"), data.get("settlement"), document_path, photo_path, data.get("remark"), now_text(), now_text()),
        )
        supplier_id = cursor.lastrowid
        if document_path:
            conn.execute(
                "INSERT INTO supplier_attachments (supplier_id,file_name,file_path,file_type,uploaded_at) VALUES (?,?,?,?,?)",
                (supplier_id, Path(document_path).name, document_path, Path(document_path).suffix.lstrip("."), now_text()),
            )
        for attachment in attachments:
            conn.execute(
                "INSERT INTO supplier_attachments (supplier_id,file_name,file_path,file_type,uploaded_at) VALUES (?,?,?,?,?)",
                (supplier_id, attachment["file_name"], attachment["file_path"], attachment["file_type"], now_text()),
            )
    log_action("新增供应商", data.get("name", ""))
    return jsonify({"ok": True})


@app.put("/api/suppliers/<int:item_id>")
@admin_required
def suppliers_update(item_id: int):
    data = payload()
    with db() as conn:
        conn.execute(
            """
            UPDATE suppliers SET name=?,level=?,contact=?,phone=?,address=?,account_info=?,company_type=?,settlement=?,remark=?,updated_at=?
            WHERE id=?
            """,
            (data.get("name"), data.get("level"), data.get("contact"), data.get("phone"), data.get("address"), data.get("account_info"), data.get("company_type"), data.get("settlement"), data.get("remark"), now_text(), item_id),
        )
    log_action("修改供应商", f"ID {item_id}")
    return jsonify({"ok": True})


@app.delete("/api/suppliers/<int:item_id>")
@admin_required
def suppliers_delete(item_id: int):
    with db() as conn:
        used = conn.execute(
            """SELECT (SELECT COUNT(*) FROM purchase_orders WHERE supplier_id=?) +
                      (SELECT COUNT(*) FROM purchase_order_headers WHERE supplier_id=?)""", (item_id, item_id)
        ).fetchone()[0]
        if used:
            return jsonify({"ok": False, "message": "该供应商已有采购单记录，不能删除；建议保留资料或后续增加停用功能"}), 400
        supplier = conn.execute("SELECT document_path,photo_path FROM suppliers WHERE id=?", (item_id,)).fetchone()
        attachment_paths = [
            row["file_path"]
            for row in conn.execute("SELECT file_path FROM supplier_attachments WHERE supplier_id=?", (item_id,))
        ]
        conn.execute("DELETE FROM supplier_attachments WHERE supplier_id=?", (item_id,))
        conn.execute("DELETE FROM suppliers WHERE id=?", (item_id,))
    if supplier:
        delete_uploaded_file(supplier["document_path"])
        delete_uploaded_file(supplier["photo_path"])
    for file_path in attachment_paths:
        delete_uploaded_file(file_path)
    log_action("删除供应商", f"ID {item_id}")
    return jsonify({"ok": True})


@app.get("/api/suppliers/<int:item_id>/attachments")
@login_required
def supplier_attachments_list(item_id: int):
    with db() as conn:
        rows = rows_to_list(conn.execute(
            "SELECT id,supplier_id,file_name,file_path,file_type,uploaded_at FROM supplier_attachments WHERE supplier_id=? ORDER BY id DESC",
            (item_id,),
        ))
    return jsonify({"ok": True, "items": rows})


@app.post("/api/suppliers/<int:item_id>/attachments")
@admin_required
def supplier_attachments_upload(item_id: int):
    try:
        attachments = save_uploads("attachments", "suppliers")
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    with db() as conn:
        supplier = conn.execute("SELECT id FROM suppliers WHERE id=?", (item_id,)).fetchone()
        if not supplier:
            return jsonify({"ok": False, "message": "供应商不存在"}), 404
        for attachment in attachments:
            conn.execute(
                "INSERT INTO supplier_attachments (supplier_id,file_name,file_path,file_type,uploaded_at) VALUES (?,?,?,?,?)",
                (item_id, attachment["file_name"], attachment["file_path"], attachment["file_type"], now_text()),
            )
    log_action("上传供应商附件", f"供应商ID {item_id}，{len(attachments)} 个文件")
    return jsonify({"ok": True, "count": len(attachments)})


@app.get("/api/attachments/<int:attachment_id>/download")
@login_required
def attachment_download(attachment_id: int):
    with db() as conn:
        attachment = conn.execute("SELECT * FROM supplier_attachments WHERE id=?", (attachment_id,)).fetchone()
    if not attachment:
        return jsonify({"ok": False, "message": "附件不存在"}), 404
    target = (WRITE_DIR / attachment["file_path"]).resolve()
    if not target.exists() or WRITE_DIR.resolve() not in target.parents:
        return jsonify({"ok": False, "message": "附件文件不存在"}), 404
    return send_file(target, as_attachment=True, download_name=attachment["file_name"])


@app.delete("/api/attachments/<int:attachment_id>")
@admin_required
def attachment_delete(attachment_id: int):
    with db() as conn:
        attachment = conn.execute("SELECT * FROM supplier_attachments WHERE id=?", (attachment_id,)).fetchone()
        if not attachment:
            return jsonify({"ok": False, "message": "附件不存在"}), 404
        conn.execute("DELETE FROM supplier_attachments WHERE id=?", (attachment_id,))
    delete_uploaded_file(attachment["file_path"])
    log_action("删除供应商附件", f"附件ID {attachment_id}")
    return jsonify({"ok": True})


@app.get("/api/customers")
@login_required
def customers_list():
    return list_table("customers", ["code", "name", "contact"])


@app.post("/api/customers")
@admin_required
def customers_create():
    data = payload()
    if not (data.get("name") or "").strip():
        return jsonify({"ok": False, "message": "客户名称必填"}), 400
    try:
        document_path = save_upload("document", "customers")
        photo_path = save_upload("photo", "customers")
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    with db() as conn:
        code = next_code(conn, "customers", "CUS")
        conn.execute(
            """
            INSERT INTO customers (code,name,level,contact,phone,address,account_info,company_type,settlement,document_path,photo_path,total_sales,received_amount,remark,created_at,updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (code, data.get("name"), data.get("level"), data.get("contact"), data.get("phone"), data.get("address"), data.get("account_info"), data.get("company_type"), data.get("settlement"), document_path, photo_path, number_value(data, "total_sales"), number_value(data, "received_amount"), data.get("remark"), now_text(), now_text()),
        )
    log_action("新增客户", data.get("name", ""))
    return jsonify({"ok": True})


@app.put("/api/customers/<int:item_id>")
@admin_required
def customers_update(item_id: int):
    data = payload()
    with db() as conn:
        conn.execute(
            """
            UPDATE customers SET name=?,level=?,contact=?,phone=?,address=?,account_info=?,company_type=?,settlement=?,total_sales=?,received_amount=?,remark=?,updated_at=?
            WHERE id=?
            """,
            (data.get("name"), data.get("level"), data.get("contact"), data.get("phone"), data.get("address"), data.get("account_info"), data.get("company_type"), data.get("settlement"), number_value(data, "total_sales"), number_value(data, "received_amount"), data.get("remark"), now_text(), item_id),
        )
    log_action("修改客户", f"ID {item_id}")
    return jsonify({"ok": True})


@app.delete("/api/customers/<int:item_id>")
@admin_required
def customers_delete(item_id: int):
    with db() as conn:
        used = conn.execute(
            """SELECT (SELECT COUNT(*) FROM sales_orders WHERE customer_id=?) +
                      (SELECT COUNT(*) FROM sales_order_headers WHERE customer_id=?)""", (item_id, item_id)
        ).fetchone()[0]
        if used:
            return jsonify({"ok": False, "message": "该客户已有销售单记录，不能删除；建议保留资料或后续增加停用功能"}), 400
        conn.execute("DELETE FROM customers WHERE id=?", (item_id,))
    log_action("删除客户", f"ID {item_id}")
    return jsonify({"ok": True})


@app.get("/api/products")
@login_required
def products_list():
    return list_table("products", ["code", "name", "equipment", "part_type"])


@app.post("/api/products")
@admin_required
def products_create():
    data = payload()
    if not (data.get("code") or "").strip() or not (data.get("name") or "").strip():
        return jsonify({"ok": False, "message": "备件编号和产品名称必填"}), 400
    try:
        photo_path = save_upload("photo", "products")
        spec_path = save_upload("spec", "products")
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    with db() as conn:
        conn.execute(
            """
            INSERT INTO products (code,name,description,unit,current_stock,safe_stock,photo_path,spec_path,equipment,part_type,warranty_until,origin,purchase_price,sale_price,remark,created_at,updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (data.get("code"), data.get("name"), data.get("description"), data.get("unit"), number_value(data, "current_stock"), number_value(data, "safe_stock"), photo_path, spec_path, data.get("equipment"), data.get("part_type"), data.get("warranty_until"), data.get("origin"), number_value(data, "purchase_price"), number_value(data, "sale_price"), data.get("remark"), now_text(), now_text()),
        )
    log_action("新增产品", data.get("code", ""))
    return jsonify({"ok": True})


@app.put("/api/products/<int:item_id>")
@admin_required
def products_update(item_id: int):
    data = payload()
    with db() as conn:
        conn.execute(
            """
            UPDATE products SET code=?,name=?,description=?,unit=?,safe_stock=?,equipment=?,part_type=?,warranty_until=?,origin=?,purchase_price=?,sale_price=?,remark=?,updated_at=?
            WHERE id=?
            """,
            (data.get("code"), data.get("name"), data.get("description"), data.get("unit"), number_value(data, "safe_stock"), data.get("equipment"), data.get("part_type"), data.get("warranty_until"), data.get("origin"), number_value(data, "purchase_price"), number_value(data, "sale_price"), data.get("remark"), now_text(), item_id),
        )
    log_action("修改产品", f"ID {item_id}")
    return jsonify({"ok": True})


@app.delete("/api/products/<int:item_id>")
@admin_required
def products_delete(item_id: int):
    with db() as conn:
        used = conn.execute(
            """
            SELECT
              (SELECT COUNT(*) FROM purchase_orders WHERE product_id=?) +
              (SELECT COUNT(*) FROM sales_orders WHERE product_id=?) +
              (SELECT COUNT(*) FROM purchase_order_items WHERE product_id=?) +
              (SELECT COUNT(*) FROM sales_order_items WHERE product_id=?) +
              (SELECT COUNT(*) FROM stock_movements WHERE product_id=?)
            """,
            (item_id, item_id, item_id, item_id, item_id),
        ).fetchone()[0]
        if used:
            return jsonify({"ok": False, "message": "该产品已有订单或库存流水，不能删除；建议保留资料或后续增加停用功能"}), 400
        conn.execute("DELETE FROM products WHERE id=?", (item_id,))
    log_action("删除产品", f"ID {item_id}")
    return jsonify({"ok": True})


@app.get("/api/template-info")
@login_required
def template_info():
    """给前端使用的模板选项，来源于 Excel 需求表。"""
    return jsonify({
        "ok": True,
        "levels": ["A", "B", "C"],
        "company_types": ["外企", "私企", "个人"],
        "settlements": ["预付100%", "预付30%，发货70%", "现金", "月结"],
        "units": ["个", "kg", "件", "套"],
        "equipments": ["熔炼", "加热", "其它"],
        "part_types": ["电气", "机械", "液体", "液压", "其它"],
        "origins": ["中国", "德国", "美国"],
    })


@app.get("/api/legacy/purchases")
@login_required
def purchases_list():
    with db() as conn:
        rows = rows_to_list(conn.execute(
            """
            SELECT po.*, s.name AS supplier_name, p.code AS product_code, p.name AS product_name,
                   p.description, p.unit, p.photo_path, p.equipment, p.part_type, p.warranty_until, p.origin
            FROM purchase_orders po
            JOIN suppliers s ON s.id=po.supplier_id
            JOIN products p ON p.id=po.product_id
            ORDER BY po.id DESC
            """
        ))
    return jsonify({"ok": True, "items": rows})


@app.post("/api/legacy/purchases")
@login_required
def purchases_create():
    data = payload()
    product_code = (data.get("product_code") or "").strip()
    product_name = (data.get("product_name") or data.get("name") or "").strip()
    quantity = number_value(data, "quantity")
    unit_price = number_value(data, "unit_price")
    if not product_code:
        return jsonify({"ok": False, "message": "备件编号必填"}), 400
    if not product_name:
        return jsonify({"ok": False, "message": "产品名称必填"}), 400
    if quantity <= 0:
        return jsonify({"ok": False, "message": "采购数量必须大于 0"}), 400
    try:
        photo_path = save_upload("photo", "products")
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    user = current_user()
    with db() as conn:
        supplier = find_purchase_supplier(conn, data)
        product_id = save_purchase_product(conn, data, unit_price, photo_path)
        order_no = (data.get("order_no") or "").strip() or next_order_no(conn, "purchase_orders", "PO")
        exists = conn.execute("SELECT id FROM purchase_orders WHERE order_no=?", (order_no,)).fetchone()
        if exists:
            return jsonify({"ok": False, "message": "采购单号已存在，请更换采购单号"}), 400
        order_date = data.get("order_date") or datetime.now().strftime("%Y-%m-%d")
        amount = quantity * unit_price
        conn.execute(
            """
            INSERT INTO purchase_orders (order_no,supplier_id,order_date,buyer,product_id,quantity,unit_price,amount,status,remark,created_at,updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (order_no, supplier["id"], order_date, user["display_name"], product_id, quantity, unit_price, amount, "approved", data.get("remark"), now_text(), now_text()),
        )
        conn.execute("UPDATE products SET current_stock=current_stock+?, updated_at=? WHERE id=?", (quantity, now_text(), product_id))
        conn.execute(
            "INSERT INTO stock_movements (product_id,movement_time,movement_type,quantity,operator,source_no,remark) VALUES (?,?,?,?,?,?,?)",
            (product_id, now_text(), "采购入库", quantity, user["display_name"], order_no, "采购单新增自动入库"),
        )
    log_action("新增采购单", order_no)
    return jsonify({"ok": True, "order_no": order_no})


@app.put("/api/legacy/purchases/<int:item_id>")
@login_required
def purchases_update(item_id: int):
    data = payload()
    product_code = (data.get("product_code") or "").strip()
    product_name = (data.get("product_name") or data.get("name") or "").strip()
    quantity = number_value(data, "quantity")
    unit_price = number_value(data, "unit_price")
    if not product_code:
        return jsonify({"ok": False, "message": "备件编号必填"}), 400
    if not product_name:
        return jsonify({"ok": False, "message": "产品名称必填"}), 400
    if quantity <= 0:
        return jsonify({"ok": False, "message": "采购数量必须大于 0"}), 400
    try:
        photo_path = save_upload("photo", "products")
    except ValueError as exc:
        return jsonify({"ok": False, "message": str(exc)}), 400
    user = current_user()
    with db() as conn:
        old_order = conn.execute("SELECT * FROM purchase_orders WHERE id=?", (item_id,)).fetchone()
        if not old_order:
            return jsonify({"ok": False, "message": "采购单不存在"}), 404
        supplier = find_purchase_supplier(conn, data)
        order_no = (data.get("order_no") or "").strip() or old_order["order_no"]
        exists = conn.execute("SELECT id FROM purchase_orders WHERE order_no=? AND id<>?", (order_no, item_id)).fetchone()
        if exists:
            return jsonify({"ok": False, "message": "采购单号已存在，请更换采购单号"}), 400
        order_date = data.get("order_date") or old_order["order_date"]
        product_id = save_purchase_product(conn, data, unit_price, photo_path)
        amount = quantity * unit_price
        old_product = conn.execute("SELECT current_stock FROM products WHERE id=?", (old_order["product_id"],)).fetchone()
        new_product = conn.execute("SELECT current_stock FROM products WHERE id=?", (product_id,)).fetchone()
        if product_id == old_order["product_id"]:
            stock_delta = quantity - old_order["quantity"]
            if old_product and old_product["current_stock"] + stock_delta < 0:
                return jsonify({"ok": False, "message": "修改后库存会小于 0，请先检查后续销售或出库记录"}), 400
            conn.execute("UPDATE products SET current_stock=current_stock+?, updated_at=? WHERE id=?", (stock_delta, now_text(), product_id))
            if stock_delta:
                conn.execute(
                    "INSERT INTO stock_movements (product_id,movement_time,movement_type,quantity,operator,source_no,remark) VALUES (?,?,?,?,?,?,?)",
                    (product_id, now_text(), "采购调整", stock_delta, user["display_name"], old_order["order_no"], "手动编辑采购单"),
                )
        else:
            if old_product and old_product["current_stock"] - old_order["quantity"] < 0:
                return jsonify({"ok": False, "message": "原产品库存不足以撤回这张采购单，请先检查后续销售或出库记录"}), 400
            conn.execute("UPDATE products SET current_stock=current_stock-?, updated_at=? WHERE id=?", (old_order["quantity"], now_text(), old_order["product_id"]))
            conn.execute("UPDATE products SET current_stock=current_stock+?, updated_at=? WHERE id=?", (quantity, now_text(), product_id))
            conn.execute(
                "INSERT INTO stock_movements (product_id,movement_time,movement_type,quantity,operator,source_no,remark) VALUES (?,?,?,?,?,?,?)",
                (old_order["product_id"], now_text(), "采购调整", -old_order["quantity"], user["display_name"], old_order["order_no"], "手动编辑采购单，撤回原产品入库"),
            )
            conn.execute(
                "INSERT INTO stock_movements (product_id,movement_time,movement_type,quantity,operator,source_no,remark) VALUES (?,?,?,?,?,?,?)",
                (product_id, now_text(), "采购调整", quantity, user["display_name"], old_order["order_no"], "手动编辑采购单，写入新产品入库"),
            )
        conn.execute(
            """
            UPDATE purchase_orders
            SET order_no=?,supplier_id=?,order_date=?,product_id=?,quantity=?,unit_price=?,amount=?,remark=?,updated_at=?
            WHERE id=?
            """,
            (order_no, supplier["id"], order_date, product_id, quantity, unit_price, amount, data.get("remark"), now_text(), item_id),
        )
        if order_no != old_order["order_no"]:
            conn.execute("UPDATE stock_movements SET source_no=? WHERE source_no=?", (order_no, old_order["order_no"]))
    log_action("修改采购单", f"ID {item_id}")
    return jsonify({"ok": True, "order_no": order_no})


@app.get("/api/legacy/sales")
@login_required
def sales_list():
    with db() as conn:
        rows = rows_to_list(conn.execute(
            """
            SELECT so.*, c.name AS customer_name, p.code AS product_code, p.name AS product_name,
                   p.description, p.unit
            FROM sales_orders so
            JOIN customers c ON c.id=so.customer_id
            JOIN products p ON p.id=so.product_id
            ORDER BY so.id DESC
            """
        ))
    return jsonify({"ok": True, "items": rows})


@app.post("/api/legacy/sales")
@login_required
def sales_create():
    data = payload()
    product_code = (data.get("product_code") or "").strip()
    product_name = (data.get("product_name") or data.get("name") or "").strip()
    quantity = number_value(data, "quantity")
    unit_price = number_value(data, "unit_price")
    if not product_code:
        return jsonify({"ok": False, "message": "备件编号必填"}), 400
    if not product_name:
        return jsonify({"ok": False, "message": "产品名称必填"}), 400
    if quantity <= 0:
        return jsonify({"ok": False, "message": "销售数量必须大于 0"}), 400
    user = current_user()
    with db() as conn:
        customer = get_or_create_default_customer(conn)
        product = conn.execute("SELECT * FROM products WHERE code=?", (product_code,)).fetchone()
        if not product:
            return jsonify({"ok": False, "message": "备件编号不存在，请先在产品管理或采购管理中建立该产品"}), 400
        if product["current_stock"] < quantity:
            return jsonify({"ok": False, "message": f"库存不足，当前库存为 {product['current_stock']}"}), 400
        conn.execute(
            """
            UPDATE products
            SET name=?,description=?,unit=?,sale_price=?,updated_at=?
            WHERE id=?
            """,
            (product_name, data.get("description"), data.get("unit"), unit_price, now_text(), product["id"]),
        )
        order_no = next_order_no(conn, "sales_orders", "SO")
        amount = quantity * unit_price
        conn.execute(
            """
            INSERT INTO sales_orders (order_no,customer_id,order_date,seller,product_id,quantity,unit_price,amount,delivery_date,remark,created_at,updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (order_no, customer["id"], datetime.now().strftime("%Y-%m-%d"), user["display_name"], product["id"], quantity, unit_price, amount, data.get("delivery_date"), data.get("remark"), now_text(), now_text()),
        )
        conn.execute("UPDATE products SET current_stock=current_stock-?, updated_at=? WHERE id=?", (quantity, now_text(), product["id"]))
        conn.execute("UPDATE customers SET total_sales=total_sales+?, updated_at=? WHERE id=?", (amount, now_text(), customer["id"]))
        conn.execute(
            "INSERT INTO stock_movements (product_id,movement_time,movement_type,quantity,operator,source_no,remark) VALUES (?,?,?,?,?,?,?)",
            (product["id"], now_text(), "销售出库", -quantity, user["display_name"], order_no, "销售单新增自动出库"),
        )
    log_action("新增销售单", order_no)
    return jsonify({"ok": True, "order_no": order_no})


@app.put("/api/legacy/sales/<int:item_id>")
@admin_required
def sales_update(item_id: int):
    data = payload()
    product_code = (data.get("product_code") or "").strip()
    product_name = (data.get("product_name") or data.get("name") or "").strip()
    quantity = number_value(data, "quantity")
    unit_price = number_value(data, "unit_price")
    if not product_code:
        return jsonify({"ok": False, "message": "备件编号必填"}), 400
    if not product_name:
        return jsonify({"ok": False, "message": "产品名称必填"}), 400
    if quantity <= 0:
        return jsonify({"ok": False, "message": "销售数量必须大于 0"}), 400
    user = current_user()
    with db() as conn:
        old_order = conn.execute("SELECT * FROM sales_orders WHERE id=?", (item_id,)).fetchone()
        if not old_order:
            return jsonify({"ok": False, "message": "销售单不存在"}), 404
        if old_order["status"] == "voided":
            return jsonify({"ok": False, "message": "已作废的销售单不能修改"}), 400
        product = conn.execute("SELECT * FROM products WHERE code=?", (product_code,)).fetchone()
        if not product:
            return jsonify({"ok": False, "message": "备件编号不存在，请先建立产品"}), 400
        amount = quantity * unit_price
        if product["id"] == old_order["product_id"]:
            # 先把原销售数量视为退回，再按新数量扣减，等价于库存变化 old - new。
            stock_delta = old_order["quantity"] - quantity
            if product["current_stock"] + stock_delta < 0:
                return jsonify({"ok": False, "message": "修改后库存不足，请先检查库存"}), 400
            conn.execute("UPDATE products SET current_stock=current_stock+?, name=?,description=?,unit=?,sale_price=?,updated_at=? WHERE id=?", (stock_delta, product_name, data.get("description"), data.get("unit"), unit_price, now_text(), product["id"]))
            if stock_delta:
                conn.execute(
                    "INSERT INTO stock_movements (product_id,movement_time,movement_type,quantity,operator,source_no,remark) VALUES (?,?,?,?,?,?,?)",
                    (product["id"], now_text(), "销售调整", stock_delta, user["display_name"], old_order["order_no"], "手动编辑销售单"),
                )
        else:
            if product["current_stock"] < quantity:
                return jsonify({"ok": False, "message": f"新产品库存不足，当前库存为 {product['current_stock']}"}), 400
            conn.execute("UPDATE products SET current_stock=current_stock+?, updated_at=? WHERE id=?", (old_order["quantity"], now_text(), old_order["product_id"]))
            conn.execute("UPDATE products SET current_stock=current_stock-?, name=?,description=?,unit=?,sale_price=?,updated_at=? WHERE id=?", (quantity, product_name, data.get("description"), data.get("unit"), unit_price, now_text(), product["id"]))
            conn.execute(
                "INSERT INTO stock_movements (product_id,movement_time,movement_type,quantity,operator,source_no,remark) VALUES (?,?,?,?,?,?,?)",
                (old_order["product_id"], now_text(), "销售调整", old_order["quantity"], user["display_name"], old_order["order_no"], "手动编辑销售单，回补原产品"),
            )
            conn.execute(
                "INSERT INTO stock_movements (product_id,movement_time,movement_type,quantity,operator,source_no,remark) VALUES (?,?,?,?,?,?,?)",
                (product["id"], now_text(), "销售调整", -quantity, user["display_name"], old_order["order_no"], "手动编辑销售单，扣减新产品"),
            )
        conn.execute("UPDATE customers SET total_sales=total_sales-?+?, updated_at=? WHERE id=?", (old_order["amount"], amount, now_text(), old_order["customer_id"]))
        conn.execute(
            """
            UPDATE sales_orders
            SET product_id=?,quantity=?,unit_price=?,amount=?,delivery_date=?,remark=?,updated_at=?
            WHERE id=?
            """,
            (product["id"], quantity, unit_price, amount, data.get("delivery_date"), data.get("remark"), now_text(), item_id),
        )
    log_action("修改销售单", f"ID {item_id}")
    return jsonify({"ok": True, "order_no": old_order["order_no"]})


@app.post("/api/legacy/sales/<int:item_id>/void")
@admin_required
def sales_void(item_id: int):
    user = current_user()
    with db() as conn:
        order = conn.execute("SELECT * FROM sales_orders WHERE id=?", (item_id,)).fetchone()
        if not order:
            return jsonify({"ok": False, "message": "销售单不存在"}), 404
        if order["status"] == "voided":
            return jsonify({"ok": False, "message": "销售单已经作废"}), 400
        conn.execute("UPDATE sales_orders SET status='voided', updated_at=? WHERE id=?", (now_text(), item_id))
        conn.execute("UPDATE products SET current_stock=current_stock+?, updated_at=? WHERE id=?", (order["quantity"], now_text(), order["product_id"]))
        conn.execute("UPDATE customers SET total_sales=total_sales-?, updated_at=? WHERE id=?", (order["amount"], now_text(), order["customer_id"]))
        conn.execute(
            "INSERT INTO stock_movements (product_id,movement_time,movement_type,quantity,operator,source_no,remark) VALUES (?,?,?,?,?,?,?)",
            (order["product_id"], now_text(), "销售作废回补", order["quantity"], user["display_name"], order["order_no"], "作废销售单自动回补库存"),
        )
    log_action("作废销售单", order["order_no"])
    return jsonify({"ok": True, "order_no": order["order_no"]})


def clean_order_items(conn: sqlite3.Connection, raw_items: list, price_field: str) -> tuple[list[dict], str | None]:
    """校验整单明细，并补齐产品名称、单位等显示信息。"""
    if not isinstance(raw_items, list) or not raw_items:
        return [], "请至少添加一条备件明细"
    cleaned = []
    seen = set()
    for index, item in enumerate(raw_items, start=1):
        try:
            product_id = int(item.get("product_id") or 0)
            quantity = float(item.get("quantity") or 0)
            unit_price = float(item.get("unit_price") or 0)
        except (TypeError, ValueError):
            return [], f"第 {index} 条明细的数量或单价格式不正确"
        product = conn.execute("SELECT * FROM products WHERE id=?", (product_id,)).fetchone()
        if not product:
            return [], f"第 {index} 条明细的备件不存在"
        if product_id in seen:
            return [], f"备件 {product['code']} 在同一订单中重复，请合并数量"
        if quantity <= 0:
            return [], f"第 {index} 条明细的数量必须大于 0"
        if unit_price < 0:
            return [], f"第 {index} 条明细的单价不能小于 0"
        seen.add(product_id)
        cleaned.append({
            "product_id": product_id,
            "product_code": product["code"],
            "product_name": product["name"],
            "quantity": quantity,
            "unit_price": unit_price,
            "amount": quantity * unit_price,
            "remark": (item.get("remark") or "").strip(),
            "price_field": price_field,
        })
    return cleaned, None


def order_items(conn: sqlite3.Connection, order_id: int, kind: str) -> list[dict]:
    table = "purchase_order_items" if kind == "purchase" else "sales_order_items"
    processed = "received_quantity" if kind == "purchase" else "shipped_quantity"
    return rows_to_list(conn.execute(
        f"""
        SELECT oi.*, p.code AS product_code, p.name AS product_name, p.description, p.unit,
               p.current_stock, oi.{processed} AS processed_quantity,
               oi.quantity - oi.{processed} AS remaining_quantity
        FROM {table} oi JOIN products p ON p.id=oi.product_id
        WHERE oi.order_id=? ORDER BY oi.id
        """,
        (order_id,),
    ))


@app.get("/api/purchases")
@login_required
def purchase_headers_list():
    q = (request.args.get("q") or "").strip()
    params = []
    where = ""
    if q:
        where = "WHERE poh.order_no LIKE ? OR s.name LIKE ?"
        params = [f"%{q}%", f"%{q}%"]
    with db() as conn:
        rows = rows_to_list(conn.execute(
            f"""
            SELECT poh.*, s.code AS supplier_code, s.name AS supplier_name,
                   COUNT(poi.id) AS item_count, COALESCE(SUM(poi.quantity),0) AS total_quantity,
                   COALESCE(SUM(poi.received_quantity),0) AS processed_quantity,
                   COALESCE(SUM(poi.amount),0) AS amount,
                   CASE WHEN poh.status!='completed' AND poh.expected_date!=''
                             AND date(poh.expected_date) < date('now','localtime') THEN 1 ELSE 0 END AS overdue
            FROM purchase_order_headers poh
            JOIN suppliers s ON s.id=poh.supplier_id
            LEFT JOIN purchase_order_items poi ON poi.order_id=poh.id
            {where}
            GROUP BY poh.id ORDER BY poh.id DESC
            """,
            params,
        ))
        for row in rows:
            row["items"] = order_items(conn, row["id"], "purchase")
    return jsonify({"ok": True, "items": rows})


@app.post("/api/purchases")
@login_required
def purchase_header_create():
    data = payload()
    user = current_user()
    with db() as conn:
        supplier = conn.execute("SELECT id FROM suppliers WHERE id=?", (data.get("supplier_id"),)).fetchone()
        if not supplier:
            return jsonify({"ok": False, "message": "请选择供应商"}), 400
        items, error = clean_order_items(conn, data.get("items"), "purchase_price")
        if error:
            return jsonify({"ok": False, "message": error}), 400
        order_no = (data.get("order_no") or "").strip() or next_order_no(conn, "purchase_order_headers", "PO")
        if conn.execute("SELECT 1 FROM purchase_order_headers WHERE order_no=?", (order_no,)).fetchone():
            return jsonify({"ok": False, "message": "采购单号已存在"}), 400
        order_date = data.get("order_date") or datetime.now().strftime("%Y-%m-%d")
        expected_date = data.get("expected_date") or ""
        cursor = conn.execute(
            """
            INSERT INTO purchase_order_headers
            (order_no,supplier_id,order_date,expected_date,buyer,status,remark,created_at,updated_at)
            VALUES (?,?,?,?,?,'pending',?,?,?)
            """,
            (order_no, supplier["id"], order_date, expected_date, user["display_name"],
             data.get("remark"), now_text(), now_text()),
        )
        for item in items:
            conn.execute(
                """INSERT INTO purchase_order_items
                (order_id,product_id,quantity,unit_price,amount,received_quantity,remark)
                VALUES (?,?,?,?,?,0,?)""",
                (cursor.lastrowid, item["product_id"], item["quantity"], item["unit_price"], item["amount"], item["remark"]),
            )
            conn.execute("UPDATE products SET purchase_price=?,updated_at=? WHERE id=?",
                         (item["unit_price"], now_text(), item["product_id"]))
    log_action("新增采购整单", order_no)
    return jsonify({"ok": True, "order_no": order_no})


@app.put("/api/purchases/<int:order_id>")
@admin_required
def purchase_header_update(order_id: int):
    data = payload()
    with db() as conn:
        order = conn.execute("SELECT * FROM purchase_order_headers WHERE id=?", (order_id,)).fetchone()
        if not order:
            return jsonify({"ok": False, "message": "采购单不存在"}), 404
        if order["status"] != "pending":
            return jsonify({"ok": False, "message": "已经开始到货的采购单不能修改明细"}), 400
        supplier = conn.execute("SELECT id FROM suppliers WHERE id=?", (data.get("supplier_id"),)).fetchone()
        if not supplier:
            return jsonify({"ok": False, "message": "请选择供应商"}), 400
        items, error = clean_order_items(conn, data.get("items"), "purchase_price")
        if error:
            return jsonify({"ok": False, "message": error}), 400
        order_no = (data.get("order_no") or order["order_no"]).strip()
        duplicate = conn.execute("SELECT 1 FROM purchase_order_headers WHERE order_no=? AND id<>?", (order_no, order_id)).fetchone()
        if duplicate:
            return jsonify({"ok": False, "message": "采购单号已存在"}), 400
        conn.execute(
            """UPDATE purchase_order_headers SET order_no=?,supplier_id=?,order_date=?,expected_date=?,remark=?,updated_at=? WHERE id=?""",
            (order_no, supplier["id"], data.get("order_date") or order["order_date"], data.get("expected_date") or "",
             data.get("remark"), now_text(), order_id),
        )
        conn.execute("DELETE FROM purchase_order_items WHERE order_id=?", (order_id,))
        for item in items:
            conn.execute(
                "INSERT INTO purchase_order_items (order_id,product_id,quantity,unit_price,amount,received_quantity,remark) VALUES (?,?,?,?,?,0,?)",
                (order_id, item["product_id"], item["quantity"], item["unit_price"], item["amount"], item["remark"]),
            )
    log_action("修改采购整单", order_no)
    return jsonify({"ok": True, "order_no": order_no})


@app.post("/api/purchases/<int:order_id>/receive")
@login_required
def purchase_receive(order_id: int):
    data = payload()
    entries = data.get("items") or []
    user = current_user()
    with db() as conn:
        order = conn.execute("SELECT * FROM purchase_order_headers WHERE id=?", (order_id,)).fetchone()
        if not order:
            return jsonify({"ok": False, "message": "采购单不存在"}), 404
        if order["status"] == "completed":
            return jsonify({"ok": False, "message": "采购单已全部到货，不能重复入库"}), 400
        changes = []
        for entry in entries:
            try:
                item_id = int(entry.get("item_id") or 0)
                quantity = float(entry.get("quantity") or 0)
            except (TypeError, ValueError):
                return jsonify({"ok": False, "message": "到货数量格式不正确"}), 400
            if quantity <= 0:
                continue
            item = conn.execute("SELECT * FROM purchase_order_items WHERE id=? AND order_id=?", (item_id, order_id)).fetchone()
            if not item or quantity > item["quantity"] - item["received_quantity"] + 0.000001:
                return jsonify({"ok": False, "message": "到货数量超过该备件的未到货数量"}), 400
            changes.append((item, quantity))
        if not changes:
            return jsonify({"ok": False, "message": "请填写本次实际到货数量"}), 400
        for item, quantity in changes:
            conn.execute("UPDATE purchase_order_items SET received_quantity=received_quantity+? WHERE id=?", (quantity, item["id"]))
            conn.execute("UPDATE products SET current_stock=current_stock+?,updated_at=? WHERE id=?", (quantity, now_text(), item["product_id"]))
            conn.execute(
                "INSERT INTO stock_movements (product_id,movement_time,movement_type,quantity,operator,source_no,remark) VALUES (?,?,?,?,?,?,?)",
                (item["product_id"], now_text(), "采购到货入库", quantity, user["display_name"], order["order_no"], data.get("remark") or "分批到货确认"),
            )
        remaining = conn.execute(
            "SELECT COALESCE(SUM(quantity-received_quantity),0) FROM purchase_order_items WHERE order_id=?", (order_id,)
        ).fetchone()[0]
        status = "completed" if remaining <= 0.000001 else "partial"
        conn.execute(
            "UPDATE purchase_order_headers SET status=?,updated_at=?,completed_at=? WHERE id=?",
            (status, now_text(), now_text() if status == "completed" else None, order_id),
        )
    log_action("采购到货确认", f"{order['order_no']} 状态 {status}")
    return jsonify({"ok": True, "order_no": order["order_no"], "status": status})


@app.get("/api/sales")
@login_required
def sales_headers_list():
    q = (request.args.get("q") or "").strip()
    params = []
    where = ""
    if q:
        where = "WHERE soh.order_no LIKE ? OR c.name LIKE ?"
        params = [f"%{q}%", f"%{q}%"]
    with db() as conn:
        rows = rows_to_list(conn.execute(
            f"""
            SELECT soh.*, c.code AS customer_code, c.name AS customer_name,
                   COUNT(soi.id) AS item_count, COALESCE(SUM(soi.quantity),0) AS total_quantity,
                   COALESCE(SUM(soi.shipped_quantity),0) AS processed_quantity,
                   COALESCE(SUM(soi.amount),0) AS amount,
                   CASE WHEN soh.status NOT IN ('completed','cancelled') AND soh.delivery_date!=''
                             AND date(soh.delivery_date) < date('now','localtime') THEN 1 ELSE 0 END AS overdue
            FROM sales_order_headers soh
            JOIN customers c ON c.id=soh.customer_id
            LEFT JOIN sales_order_items soi ON soi.order_id=soh.id
            {where}
            GROUP BY soh.id ORDER BY soh.id DESC
            """,
            params,
        ))
        for row in rows:
            row["items"] = order_items(conn, row["id"], "sales")
    return jsonify({"ok": True, "items": rows})


@app.post("/api/sales")
@login_required
def sales_header_create():
    data = payload()
    user = current_user()
    with db() as conn:
        customer = conn.execute("SELECT id FROM customers WHERE id=?", (data.get("customer_id"),)).fetchone()
        if not customer:
            return jsonify({"ok": False, "message": "请选择客户"}), 400
        items, error = clean_order_items(conn, data.get("items"), "sale_price")
        if error:
            return jsonify({"ok": False, "message": error}), 400
        order_no = (data.get("order_no") or "").strip() or next_order_no(conn, "sales_order_headers", "SO")
        if conn.execute("SELECT 1 FROM sales_order_headers WHERE order_no=?", (order_no,)).fetchone():
            return jsonify({"ok": False, "message": "销售单号已存在"}), 400
        order_date = data.get("order_date") or datetime.now().strftime("%Y-%m-%d")
        cursor = conn.execute(
            """
            INSERT INTO sales_order_headers
            (order_no,customer_id,order_date,delivery_date,seller,status,remark,created_at,updated_at)
            VALUES (?,?,?,?,?,'pending',?,?,?)
            """,
            (order_no, customer["id"], order_date, data.get("delivery_date") or "", user["display_name"],
             data.get("remark"), now_text(), now_text()),
        )
        total = 0
        for item in items:
            total += item["amount"]
            conn.execute(
                """INSERT INTO sales_order_items
                (order_id,product_id,quantity,unit_price,amount,shipped_quantity,remark)
                VALUES (?,?,?,?,?,0,?)""",
                (cursor.lastrowid, item["product_id"], item["quantity"], item["unit_price"], item["amount"], item["remark"]),
            )
            conn.execute("UPDATE products SET sale_price=?,updated_at=? WHERE id=?", (item["unit_price"], now_text(), item["product_id"]))
        conn.execute("UPDATE customers SET total_sales=total_sales+?,updated_at=? WHERE id=?", (total, now_text(), customer["id"]))
    log_action("新增销售整单", order_no)
    return jsonify({"ok": True, "order_no": order_no})


@app.put("/api/sales/<int:order_id>")
@admin_required
def sales_header_update(order_id: int):
    data = payload()
    with db() as conn:
        order = conn.execute("SELECT * FROM sales_order_headers WHERE id=?", (order_id,)).fetchone()
        if not order:
            return jsonify({"ok": False, "message": "销售单不存在"}), 404
        if order["status"] != "pending":
            return jsonify({"ok": False, "message": "已经开始发货的销售单不能修改明细"}), 400
        customer = conn.execute("SELECT id FROM customers WHERE id=?", (data.get("customer_id"),)).fetchone()
        if not customer:
            return jsonify({"ok": False, "message": "请选择客户"}), 400
        items, error = clean_order_items(conn, data.get("items"), "sale_price")
        if error:
            return jsonify({"ok": False, "message": error}), 400
        order_no = (data.get("order_no") or order["order_no"]).strip()
        if conn.execute("SELECT 1 FROM sales_order_headers WHERE order_no=? AND id<>?", (order_no, order_id)).fetchone():
            return jsonify({"ok": False, "message": "销售单号已存在"}), 400
        old_total = conn.execute("SELECT COALESCE(SUM(amount),0) FROM sales_order_items WHERE order_id=?", (order_id,)).fetchone()[0]
        new_total = sum(item["amount"] for item in items)
        conn.execute("UPDATE customers SET total_sales=total_sales-?,updated_at=? WHERE id=?", (old_total, now_text(), order["customer_id"]))
        conn.execute("UPDATE customers SET total_sales=total_sales+?,updated_at=? WHERE id=?", (new_total, now_text(), customer["id"]))
        conn.execute(
            """UPDATE sales_order_headers SET order_no=?,customer_id=?,order_date=?,delivery_date=?,remark=?,updated_at=? WHERE id=?""",
            (order_no, customer["id"], data.get("order_date") or order["order_date"], data.get("delivery_date") or "",
             data.get("remark"), now_text(), order_id),
        )
        conn.execute("DELETE FROM sales_order_items WHERE order_id=?", (order_id,))
        for item in items:
            conn.execute(
                "INSERT INTO sales_order_items (order_id,product_id,quantity,unit_price,amount,shipped_quantity,remark) VALUES (?,?,?,?,?,0,?)",
                (order_id, item["product_id"], item["quantity"], item["unit_price"], item["amount"], item["remark"]),
            )
    log_action("修改销售整单", order_no)
    return jsonify({"ok": True, "order_no": order_no})


@app.post("/api/sales/<int:order_id>/ship")
@login_required
def sales_ship(order_id: int):
    data = payload()
    entries = data.get("items") or []
    user = current_user()
    with db() as conn:
        order = conn.execute("SELECT * FROM sales_order_headers WHERE id=?", (order_id,)).fetchone()
        if not order:
            return jsonify({"ok": False, "message": "销售单不存在"}), 404
        if order["status"] == "completed":
            return jsonify({"ok": False, "message": "销售单已全部发货，不能重复出库"}), 400
        changes = []
        stock_needed = {}
        for entry in entries:
            try:
                item_id = int(entry.get("item_id") or 0)
                quantity = float(entry.get("quantity") or 0)
            except (TypeError, ValueError):
                return jsonify({"ok": False, "message": "发货数量格式不正确"}), 400
            if quantity <= 0:
                continue
            item = conn.execute("SELECT * FROM sales_order_items WHERE id=? AND order_id=?", (item_id, order_id)).fetchone()
            if not item or quantity > item["quantity"] - item["shipped_quantity"] + 0.000001:
                return jsonify({"ok": False, "message": "发货数量超过该备件的未发货数量"}), 400
            stock_needed[item["product_id"]] = stock_needed.get(item["product_id"], 0) + quantity
            changes.append((item, quantity))
        if not changes:
            return jsonify({"ok": False, "message": "请填写本次实际发货数量"}), 400
        for product_id, needed in stock_needed.items():
            product = conn.execute("SELECT code,current_stock FROM products WHERE id=?", (product_id,)).fetchone()
            if not product or product["current_stock"] + 0.000001 < needed:
                current = product["current_stock"] if product else 0
                code = product["code"] if product else str(product_id)
                return jsonify({"ok": False, "message": f"备件 {code} 库存不足，当前库存 {current}"}), 400
        for item, quantity in changes:
            conn.execute("UPDATE sales_order_items SET shipped_quantity=shipped_quantity+? WHERE id=?", (quantity, item["id"]))
            conn.execute("UPDATE products SET current_stock=current_stock-?,updated_at=? WHERE id=?", (quantity, now_text(), item["product_id"]))
            conn.execute(
                "INSERT INTO stock_movements (product_id,movement_time,movement_type,quantity,operator,source_no,remark) VALUES (?,?,?,?,?,?,?)",
                (item["product_id"], now_text(), "销售发货出库", -quantity, user["display_name"], order["order_no"], data.get("remark") or "分批发货确认"),
            )
        remaining = conn.execute(
            "SELECT COALESCE(SUM(quantity-shipped_quantity),0) FROM sales_order_items WHERE order_id=?", (order_id,)
        ).fetchone()[0]
        status = "completed" if remaining <= 0.000001 else "partial"
        conn.execute(
            "UPDATE sales_order_headers SET status=?,updated_at=?,completed_at=? WHERE id=?",
            (status, now_text(), now_text() if status == "completed" else None, order_id),
        )
    log_action("销售发货确认", f"{order['order_no']} 状态 {status}")
    return jsonify({"ok": True, "order_no": order["order_no"], "status": status})


@app.get("/api/stock")
@login_required
def stock_list():
    q = (request.args.get("q") or "").strip()
    params = []
    where = ""
    if q:
        where = "WHERE code LIKE ? OR name LIKE ? OR equipment LIKE ? OR part_type LIKE ?"
        params = [f"%{q}%"] * 4
    with db() as conn:
        products = rows_to_list(conn.execute(f"SELECT * FROM products {where} ORDER BY name", params))
        movements = rows_to_list(conn.execute(
            """
            SELECT sm.*, p.code AS product_code, p.name AS product_name
            FROM stock_movements sm JOIN products p ON p.id=sm.product_id
            ORDER BY sm.id DESC LIMIT 100
            """
        ))
    return jsonify({"ok": True, "items": products, "movements": movements})


@app.post("/api/stock/manual")
@admin_required
def stock_manual():
    data = payload()
    quantity = number_value(data, "quantity")
    if quantity == 0:
        return jsonify({"ok": False, "message": "数量不能为 0"}), 400
    movement_type = data.get("movement_type") or "手工入库"
    signed_quantity = abs(quantity) if movement_type == "手工入库" else -abs(quantity)
    user = current_user()
    with db() as conn:
        product = conn.execute("SELECT current_stock FROM products WHERE id=?", (int(data["product_id"]),)).fetchone()
        if movement_type == "手工出库" and product["current_stock"] < abs(quantity):
            return jsonify({"ok": False, "message": "库存不足，不能手工出库"}), 400
        conn.execute("UPDATE products SET current_stock=current_stock+?, updated_at=? WHERE id=?", (signed_quantity, now_text(), int(data["product_id"])))
        conn.execute(
            "INSERT INTO stock_movements (product_id,movement_time,movement_type,quantity,operator,remark) VALUES (?,?,?,?,?,?)",
            (int(data["product_id"]), now_text(), movement_type, signed_quantity, user["display_name"], data.get("remark")),
        )
    log_action(movement_type, f"产品ID {data.get('product_id')} 数量 {signed_quantity}")
    return jsonify({"ok": True})


@app.get("/api/legacy/stats")
@login_required
def stats():
    start, end = date_range_from_request()
    product_id = request.args.get("product_id")
    customer_id = request.args.get("customer_id")
    supplier_id = request.args.get("supplier_id")
    order_no = (request.args.get("order_no") or "").strip()
    purchase_where = ["order_date BETWEEN ? AND ?"]
    purchase_params: list = [start, end]
    sales_where = ["order_date BETWEEN ? AND ?"]
    sales_params: list = [start, end]
    if product_id:
        purchase_where.append("product_id=?")
        purchase_params.append(product_id)
        sales_where.append("product_id=?")
        sales_params.append(product_id)
    if supplier_id:
        purchase_where.append("supplier_id=?")
        purchase_params.append(supplier_id)
    if customer_id:
        sales_where.append("customer_id=?")
        sales_params.append(customer_id)
    if order_no:
        purchase_where.append("order_no LIKE ?")
        purchase_params.append(f"%{order_no}%")
        sales_where.append("order_no LIKE ?")
        sales_params.append(f"%{order_no}%")
    purchase_sql = " AND ".join(purchase_where)
    sales_sql = " AND ".join(sales_where)
    with db() as conn:
        purchase = row_to_dict(conn.execute(
            f"SELECT COALESCE(SUM(amount),0) amount, COALESCE(SUM(quantity),0) quantity, COUNT(*) count FROM purchase_orders WHERE {purchase_sql}",
            purchase_params,
        ).fetchone())
        sales = row_to_dict(conn.execute(
            f"SELECT COALESCE(SUM(amount),0) amount, COALESCE(SUM(quantity),0) quantity, COUNT(*) count FROM sales_orders WHERE {sales_sql}",
            sales_params,
        ).fetchone())
        customer_rank = rows_to_list(conn.execute(
            """
            SELECT c.name, COALESCE(SUM(so.amount),0) amount
            FROM customers c LEFT JOIN sales_orders so ON so.customer_id=c.id AND so.order_date BETWEEN ? AND ?
            GROUP BY c.id ORDER BY amount DESC LIMIT 10
            """,
            (start, end),
        ))
        supplier_rank = rows_to_list(conn.execute(
            """
            SELECT s.name, COALESCE(SUM(po.amount),0) amount
            FROM suppliers s LEFT JOIN purchase_orders po ON po.supplier_id=s.id AND po.order_date BETWEEN ? AND ?
            GROUP BY s.id ORDER BY amount DESC LIMIT 10
            """,
            (start, end),
        ))
        product_rank = rows_to_list(conn.execute(
            """
            SELECT p.code, p.name, p.current_stock,
            COALESCE((SELECT SUM(quantity) FROM purchase_orders WHERE product_id=p.id AND order_date BETWEEN ? AND ?),0) AS purchase_qty,
            COALESCE((SELECT SUM(quantity) FROM sales_orders WHERE product_id=p.id AND order_date BETWEEN ? AND ?),0) AS sales_qty
            FROM products p ORDER BY sales_qty DESC LIMIT 20
            """,
            (start, end, start, end),
        ))
        order_profit = rows_to_list(conn.execute(
            """
            SELECT so.order_no, so.order_date, c.name AS customer_name, p.code AS product_code, p.name AS product_name,
                   so.quantity AS sales_qty, so.amount AS sales_amount,
                   COALESCE((
                       SELECT AVG(po.unit_price) FROM purchase_orders po WHERE po.product_id = so.product_id
                   ), p.purchase_price, 0) * so.quantity AS estimated_cost,
                   so.amount - COALESCE((
                       SELECT AVG(po.unit_price) FROM purchase_orders po WHERE po.product_id = so.product_id
                   ), p.purchase_price, 0) * so.quantity AS profit
            FROM sales_orders so
            JOIN customers c ON c.id=so.customer_id
            JOIN products p ON p.id=so.product_id
            WHERE so.order_date BETWEEN ? AND ?
            ORDER BY so.id DESC
            LIMIT 50
            """,
            (start, end),
        ))
    return jsonify({
        "ok": True,
        "start": start,
        "end": end,
        "purchase": purchase,
        "sales": sales,
        "profit": (sales["amount"] if sales else 0) - (purchase["amount"] if purchase else 0),
        "customer_rank": customer_rank,
        "supplier_rank": supplier_rank,
        "product_rank": product_rank,
        "order_profit": order_profit,
    })


@app.get("/api/stats")
@login_required
def order_stats():
    """按整单结构统计，次数按订单数计算，数量和金额按明细合计。"""
    start, end = date_range_from_request()
    product_id = request.args.get("product_id")
    customer_id = request.args.get("customer_id")
    supplier_id = request.args.get("supplier_id")
    order_no = (request.args.get("order_no") or "").strip()
    purchase_filters = ["poh.order_date BETWEEN ? AND ?"]
    purchase_params: list = [start, end]
    sales_filters = ["soh.order_date BETWEEN ? AND ?", "soh.status!='cancelled'"]
    sales_params: list = [start, end]
    if product_id:
        purchase_filters.append("poi.product_id=?")
        purchase_params.append(product_id)
        sales_filters.append("soi.product_id=?")
        sales_params.append(product_id)
    if supplier_id:
        purchase_filters.append("poh.supplier_id=?")
        purchase_params.append(supplier_id)
    if customer_id:
        sales_filters.append("soh.customer_id=?")
        sales_params.append(customer_id)
    if order_no:
        purchase_filters.append("poh.order_no LIKE ?")
        purchase_params.append(f"%{order_no}%")
        sales_filters.append("soh.order_no LIKE ?")
        sales_params.append(f"%{order_no}%")
    purchase_sql = " AND ".join(purchase_filters)
    sales_sql = " AND ".join(sales_filters)
    with db() as conn:
        purchase = row_to_dict(conn.execute(
            f"""SELECT COALESCE(SUM(poi.amount),0) amount,COALESCE(SUM(poi.quantity),0) quantity,
                       COUNT(DISTINCT poh.id) count FROM purchase_order_headers poh
                 JOIN purchase_order_items poi ON poi.order_id=poh.id WHERE {purchase_sql}""", purchase_params
        ).fetchone())
        sales = row_to_dict(conn.execute(
            f"""SELECT COALESCE(SUM(soi.amount),0) amount,COALESCE(SUM(soi.quantity),0) quantity,
                       COUNT(DISTINCT soh.id) count FROM sales_order_headers soh
                 JOIN sales_order_items soi ON soi.order_id=soh.id WHERE {sales_sql}""", sales_params
        ).fetchone())
        customer_rank = rows_to_list(conn.execute(
            """SELECT c.name,COALESCE(SUM(soi.amount),0) amount FROM customers c
               LEFT JOIN sales_order_headers soh ON soh.customer_id=c.id AND soh.order_date BETWEEN ? AND ? AND soh.status!='cancelled'
               LEFT JOIN sales_order_items soi ON soi.order_id=soh.id GROUP BY c.id ORDER BY amount DESC LIMIT 10""", (start, end)
        ))
        supplier_rank = rows_to_list(conn.execute(
            """SELECT s.name,COALESCE(SUM(poi.amount),0) amount FROM suppliers s
               LEFT JOIN purchase_order_headers poh ON poh.supplier_id=s.id AND poh.order_date BETWEEN ? AND ?
               LEFT JOIN purchase_order_items poi ON poi.order_id=poh.id GROUP BY s.id ORDER BY amount DESC LIMIT 10""", (start, end)
        ))
        product_rank = rows_to_list(conn.execute(
            """SELECT p.code,p.name,p.current_stock,
               COALESCE((SELECT SUM(poi.quantity) FROM purchase_order_items poi JOIN purchase_order_headers poh ON poh.id=poi.order_id WHERE poi.product_id=p.id AND poh.order_date BETWEEN ? AND ?),0) purchase_qty,
               COALESCE((SELECT SUM(soi.quantity) FROM sales_order_items soi JOIN sales_order_headers soh ON soh.id=soi.order_id WHERE soi.product_id=p.id AND soh.order_date BETWEEN ? AND ? AND soh.status!='cancelled'),0) sales_qty
               FROM products p ORDER BY sales_qty DESC LIMIT 20""", (start, end, start, end)
        ))
        order_profit = rows_to_list(conn.execute(
            """SELECT soh.order_no,soh.order_date,c.name customer_name,p.code product_code,p.name product_name,
               soi.quantity sales_qty,soi.amount sales_amount,
               COALESCE(p.purchase_price,0)*soi.quantity estimated_cost,
               soi.amount-COALESCE(p.purchase_price,0)*soi.quantity profit
               FROM sales_order_headers soh JOIN sales_order_items soi ON soi.order_id=soh.id
               JOIN customers c ON c.id=soh.customer_id JOIN products p ON p.id=soi.product_id
               WHERE soh.order_date BETWEEN ? AND ? AND soh.status!='cancelled' ORDER BY soh.id DESC LIMIT 50""", (start, end)
        ))
    return jsonify({"ok": True, "start": start, "end": end, "purchase": purchase, "sales": sales,
                    "profit": sales["amount"] - purchase["amount"], "customer_rank": customer_rank,
                    "supplier_rank": supplier_rank, "product_rank": product_rank, "order_profit": order_profit})


@app.get("/api/logs")
@admin_required
def logs():
    with db() as conn:
        items = rows_to_list(conn.execute("SELECT * FROM system_logs ORDER BY id DESC LIMIT 200"))
    return jsonify({"ok": True, "items": items})


@app.post("/api/backup")
@admin_required
def backup():
    ensure_dirs()
    target = BACKUP_DIR / f"app_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    shutil.copy2(DB_PATH, target)
    log_action("手动备份", target.name)
    return jsonify({"ok": True, "file": str(target)})


@app.get("/api/export/<table>")
@login_required
def export_table(table: str):
    allowed = {
        "suppliers": "suppliers",
        "customers": "customers",
        "products": "products",
        "stock": "stock_movements",
    }
    if table not in allowed and table not in ("purchases", "sales"):
        return jsonify({"ok": False, "message": "不支持导出该数据"}), 400
    with db() as conn:
        if table == "purchases":
            rows = rows_to_list(conn.execute(
                """SELECT poh.order_no AS 采购单号,s.name AS 供应商,poh.order_date AS 采购日期,
                   poh.expected_date AS 要求到货日期,p.code AS 备件编号,p.name AS 产品名称,p.description AS 技术描述,
                   p.unit AS 单位,poi.quantity AS 数量,poi.received_quantity AS 已到货数量,
                   poi.unit_price AS 单价,poi.amount AS 金额,poh.status AS 状态,poh.remark AS 订单备注
                   FROM purchase_order_headers poh JOIN suppliers s ON s.id=poh.supplier_id
                   JOIN purchase_order_items poi ON poi.order_id=poh.id JOIN products p ON p.id=poi.product_id
                   ORDER BY poh.id DESC,poi.id"""
            ))
        elif table == "sales":
            rows = rows_to_list(conn.execute(
                """SELECT soh.order_no AS 销售单号,c.name AS 客户,soh.order_date AS 销售日期,
                   soh.delivery_date AS 要求发货日期,p.code AS 备件编号,p.name AS 产品名称,p.description AS 技术描述,
                   p.unit AS 单位,soi.quantity AS 数量,soi.shipped_quantity AS 已发货数量,
                   soi.unit_price AS 单价,soi.amount AS 金额,soh.status AS 状态,soh.remark AS 订单备注
                   FROM sales_order_headers soh JOIN customers c ON c.id=soh.customer_id
                   JOIN sales_order_items soi ON soi.order_id=soh.id JOIN products p ON p.id=soi.product_id
                   ORDER BY soh.id DESC,soi.id"""
            ))
        else:
            if table not in allowed:
                return jsonify({"ok": False, "message": "不支持导出该数据"}), 400
            rows = rows_to_list(conn.execute(f"SELECT * FROM {allowed[table]} ORDER BY id DESC"))
    output = io.StringIO()
    if rows:
        writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    data = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    return send_file(data, as_attachment=True, download_name=f"{table}_{today_code_date()}.csv", mimetype="text/csv")


@app.post("/api/import/<table>")
@admin_required
def import_table(table: str):
    if table not in IMPORT_FIELDS:
        return jsonify({"ok": False, "message": "当前只支持供应商、客户、产品、采购订单、销售订单导入"}), 400
    file = request.files.get("file")
    if not file:
        return jsonify({"ok": False, "message": "请选择导入文件"}), 400
    ext = file.filename.rsplit(".", 1)[-1].lower()
    rows: list[dict] = []
    if ext == "csv":
        text = file.read().decode("utf-8-sig")
        rows = list(csv.DictReader(io.StringIO(text)))
    elif ext in ("xlsx", "xlsm"):
        try:
            rows = read_excel_template_rows(file, table)
        except RuntimeError as exc:
            return jsonify({"ok": False, "message": str(exc)}), 400
    else:
        return jsonify({"ok": False, "message": "请上传 CSV 或 XLSX 文件"}), 400

    fields = IMPORT_FIELDS[table]
    if table in ("purchases", "sales"):
        return import_order_rows(table, rows, fields)
    imported = 0
    skipped: list[str] = []
    imported_rows: list[str] = []
    with db() as conn:
        for row_number, item in enumerate(rows, start=1):
            data = {field: import_value(item, field) for field in fields}
            if table == "suppliers":
                name = import_value(item, "name")
                if not name:
                    continue
                data["name"] = name
                code = import_value(item, "code") or next_code(conn, table, "SUP")
                conn.execute(
                    """
                    INSERT INTO suppliers (code,name,level,contact,phone,address,account_info,company_type,settlement,remark,created_at,updated_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (code, data["name"], data.get("level"), data.get("contact"), data.get("phone"), data.get("address"), data.get("account_info"), data.get("company_type"), data.get("settlement"), data.get("remark"), now_text(), now_text()),
                )
            elif table == "customers":
                name = import_value(item, "name")
                if not name:
                    continue
                data["name"] = name
                code = import_value(item, "code") or next_code(conn, table, "CUS")
                conn.execute(
                    """
                    INSERT INTO customers (code,name,level,contact,phone,address,account_info,company_type,settlement,total_sales,received_amount,remark,created_at,updated_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (code, data["name"], data.get("level"), data.get("contact"), data.get("phone"), data.get("address"), data.get("account_info"), data.get("company_type"), data.get("settlement"), float(data.get("total_sales") or 0), float(data.get("received_amount") or 0), data.get("remark"), now_text(), now_text()),
                )
            elif table == "products":
                name = import_value(item, "name")
                if not name:
                    continue
                data["name"] = name
                code = import_value(item, "code")
                if not code:
                    continue
                conn.execute(
                    """
                    INSERT INTO products (code,name,description,unit,current_stock,safe_stock,equipment,part_type,warranty_until,origin,purchase_price,sale_price,remark,created_at,updated_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (code, data["name"], data.get("description"), data.get("unit"), float(data.get("current_stock") or 0), float(data.get("safe_stock") or 0), data.get("equipment"), data.get("part_type"), data.get("warranty_until"), data.get("origin"), float(data.get("purchase_price") or 0), float(data.get("sale_price") or 0), data.get("remark"), now_text(), now_text()),
                )
            elif table == "purchases":
                data["supplier_id"] = request.form.get("supplier_id") or ""
                product_code = data.get("product_code")
                quantity = float(data.get("quantity") or 0)
                unit_price = float(data.get("unit_price") or 0)
                if not product_code or quantity <= 0:
                    skipped.append(f"第{row_number}行：缺少备件编号或数量")
                    continue
                supplier = find_purchase_supplier(conn, data)
                if not supplier:
                    skipped.append(f"第{row_number}行：找不到供应商")
                    continue
                product = conn.execute("SELECT id FROM products WHERE code=?", (product_code,)).fetchone()
                if not product:
                    product_name = data.get("name") or product_code
                    conn.execute(
                        """
                        INSERT INTO products (code,name,description,unit,current_stock,safe_stock,equipment,part_type,warranty_until,origin,purchase_price,sale_price,remark,created_at,updated_at)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        """,
                        (product_code, product_name, data.get("description"), data.get("unit"), 0, 0, data.get("equipment"), data.get("part_type"), data.get("warranty_until"), data.get("origin"), unit_price, 0, "采购订单导入自动建立", now_text(), now_text()),
                    )
                    product = conn.execute("SELECT id FROM products WHERE code=?", (product_code,)).fetchone()
                order_no = (data.get("order_no") or "").strip() or next_order_no(conn, "purchase_orders", "PO")
                if conn.execute("SELECT id FROM purchase_orders WHERE order_no=?", (order_no,)).fetchone():
                    skipped.append(f"第{row_number}行：采购单号已存在 {order_no}")
                    continue
                order_date = data.get("order_date") or datetime.now().strftime("%Y-%m-%d")
                amount = quantity * unit_price
                operator = current_user()["display_name"]
                conn.execute(
                    """
                    INSERT INTO purchase_orders (order_no,supplier_id,order_date,buyer,product_id,quantity,unit_price,amount,status,remark,created_at,updated_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (order_no, supplier["id"], order_date, operator, product["id"], quantity, unit_price, amount, "approved", data.get("remark"), now_text(), now_text()),
                )
                conn.execute("UPDATE products SET current_stock=current_stock+?, updated_at=? WHERE id=?", (quantity, now_text(), product["id"]))
                conn.execute(
                    "INSERT INTO stock_movements (product_id,movement_time,movement_type,quantity,operator,source_no,remark) VALUES (?,?,?,?,?,?,?)",
                    (product["id"], now_text(), "采购入库", quantity, operator, order_no, "Excel导入采购单"),
                )
            elif table == "sales":
                customer_code = data.get("customer_code")
                customer_id = request.form.get("customer_id") or ""
                product_code = data.get("product_code")
                quantity = float(data.get("quantity") or 0)
                unit_price = float(data.get("unit_price") or 0)
                if not product_code or quantity <= 0:
                    skipped.append(f"第{row_number}行：缺少备件编号或数量")
                    continue
                if customer_code:
                    customer = conn.execute("SELECT id FROM customers WHERE code=?", (customer_code,)).fetchone()
                elif customer_id:
                    customer = conn.execute("SELECT id FROM customers WHERE id=?", (customer_id,)).fetchone()
                else:
                    customer = get_or_create_default_customer(conn)
                product = conn.execute("SELECT id,current_stock FROM products WHERE code=?", (product_code,)).fetchone()
                if not customer:
                    skipped.append(f"第{row_number}行：找不到客户")
                    continue
                if not product:
                    skipped.append(f"第{row_number}行：找不到备件 {product_code}")
                    continue
                if product["current_stock"] < quantity:
                    skipped.append(f"第{row_number}行：库存不足，当前库存 {product['current_stock']}")
                    continue
                order_no = next_order_no(conn, "sales_orders", "SO")
                amount = quantity * unit_price
                operator = current_user()["display_name"]
                conn.execute(
                    """
                    INSERT INTO sales_orders (order_no,customer_id,order_date,seller,product_id,quantity,unit_price,amount,delivery_date,remark,created_at,updated_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (order_no, customer["id"], datetime.now().strftime("%Y-%m-%d"), operator, product["id"], quantity, unit_price, amount, data.get("delivery_date"), data.get("remark"), now_text(), now_text()),
                )
                conn.execute("UPDATE products SET current_stock=current_stock-?, updated_at=? WHERE id=?", (quantity, now_text(), product["id"]))
                conn.execute("UPDATE customers SET total_sales=total_sales+?, updated_at=? WHERE id=?", (amount, now_text(), customer["id"]))
                conn.execute(
                    "INSERT INTO stock_movements (product_id,movement_time,movement_type,quantity,operator,source_no,remark) VALUES (?,?,?,?,?,?,?)",
                    (product["id"], now_text(), "销售出库", -quantity, operator, order_no, "Excel导入销售单"),
                )
            imported += 1
            imported_rows.append(f"第{row_number}行")
    log_action("导入数据", f"{table} 导入 {imported} 条，跳过 {len(skipped)} 条")
    return jsonify({
        "ok": True,
        "count": imported,
        "imported_rows": imported_rows[:30],
        "imported_rows_count": len(imported_rows),
        "skipped": skipped[:30],
        "skipped_count": len(skipped),
    })


def import_order_rows(table: str, rows: list[dict], fields: list[str]):
    """把 Excel 多行按订单号合并为整单；导入不直接改变库存。"""
    imported = 0
    skipped: list[str] = []
    imported_rows: list[str] = []
    operator = current_user()["display_name"]
    with db() as conn:
        header_table = "purchase_order_headers" if table == "purchases" else "sales_order_headers"
        prefix = "PO" if table == "purchases" else "SO"
        fallback_order_no = next_order_no(conn, header_table, prefix)
        created_orders: dict[str, int] = {}
        for row_number, raw in enumerate(rows, start=1):
            data = {field: import_value(raw, field) for field in fields}
            order_no = (data.get("order_no") or "").strip() or fallback_order_no
            product_code = (data.get("product_code") or "").strip()
            try:
                quantity = float(data.get("quantity") or 0)
                unit_price = float(data.get("unit_price") or 0)
            except (TypeError, ValueError):
                skipped.append(f"第{row_number}行：数量或单价格式不正确")
                continue
            if not product_code or quantity <= 0:
                skipped.append(f"第{row_number}行：缺少备件编号或数量")
                continue
            product = conn.execute("SELECT * FROM products WHERE code=?", (product_code,)).fetchone()
            if not product and table == "purchases":
                product_name = data.get("name") or product_code
                cursor = conn.execute(
                    """INSERT INTO products
                    (code,name,description,unit,current_stock,safe_stock,equipment,part_type,warranty_until,origin,purchase_price,sale_price,remark,created_at,updated_at)
                    VALUES (?,?,?,?,0,0,?,?,?,?,?,0,?,?,?)""",
                    (product_code, product_name, data.get("description"), data.get("unit"), data.get("equipment"),
                     data.get("part_type"), data.get("warranty_until"), data.get("origin"), unit_price,
                     "采购订单导入自动建立", now_text(), now_text()),
                )
                product = conn.execute("SELECT * FROM products WHERE id=?", (cursor.lastrowid,)).fetchone()
            if not product:
                skipped.append(f"第{row_number}行：找不到备件 {product_code}")
                continue

            order_id = created_orders.get(order_no)
            if not order_id:
                if conn.execute(f"SELECT 1 FROM {header_table} WHERE order_no=?", (order_no,)).fetchone():
                    skipped.append(f"第{row_number}行：订单号已存在 {order_no}")
                    continue
                order_date = data.get("order_date") or datetime.now().strftime("%Y-%m-%d")
                if table == "purchases":
                    data["supplier_id"] = request.form.get("supplier_id") or ""
                    supplier = find_purchase_supplier(conn, data)
                    cursor = conn.execute(
                        """INSERT INTO purchase_order_headers
                        (order_no,supplier_id,order_date,expected_date,buyer,status,remark,created_at,updated_at)
                        VALUES (?,?,?,?,?,'pending',?,?,?)""",
                        (order_no, supplier["id"], order_date, data.get("expected_date") or "", operator,
                         data.get("remark"), now_text(), now_text()),
                    )
                else:
                    customer_id = request.form.get("customer_id") or ""
                    customer = conn.execute("SELECT id FROM customers WHERE id=?", (customer_id,)).fetchone() if customer_id else None
                    if not customer and data.get("customer_code"):
                        customer = conn.execute("SELECT id FROM customers WHERE code=?", (data["customer_code"],)).fetchone()
                    customer = customer or get_or_create_default_customer(conn)
                    cursor = conn.execute(
                        """INSERT INTO sales_order_headers
                        (order_no,customer_id,order_date,delivery_date,seller,status,remark,created_at,updated_at)
                        VALUES (?,?,?,?,?,'pending',?,?,?)""",
                        (order_no, customer["id"], order_date, data.get("delivery_date") or "", operator,
                         data.get("remark"), now_text(), now_text()),
                    )
                order_id = cursor.lastrowid
                created_orders[order_no] = order_id

            item_table = "purchase_order_items" if table == "purchases" else "sales_order_items"
            processed_field = "received_quantity" if table == "purchases" else "shipped_quantity"
            existing_item = conn.execute(
                f"SELECT id,quantity FROM {item_table} WHERE order_id=? AND product_id=?", (order_id, product["id"])
            ).fetchone()
            if existing_item:
                new_quantity = existing_item["quantity"] + quantity
                conn.execute(
                    f"UPDATE {item_table} SET quantity=?,unit_price=?,amount=? WHERE id=?",
                    (new_quantity, unit_price, new_quantity * unit_price, existing_item["id"]),
                )
            else:
                conn.execute(
                    f"""INSERT INTO {item_table}
                    (order_id,product_id,quantity,unit_price,amount,{processed_field},remark)
                    VALUES (?,?,?,?,?,0,?)""",
                    (order_id, product["id"], quantity, unit_price, quantity * unit_price, data.get("remark")),
                )
            price_field = "purchase_price" if table == "purchases" else "sale_price"
            conn.execute(f"UPDATE products SET {price_field}=?,updated_at=? WHERE id=?", (unit_price, now_text(), product["id"]))
            if table == "sales":
                customer_id = conn.execute("SELECT customer_id FROM sales_order_headers WHERE id=?", (order_id,)).fetchone()[0]
                conn.execute("UPDATE customers SET total_sales=total_sales+?,updated_at=? WHERE id=?",
                             (quantity * unit_price, now_text(), customer_id))
            imported += 1
            imported_rows.append(f"第{row_number}行")
    log_action("导入整单", f"{table} 导入 {imported} 行，形成 {len(created_orders)} 张订单")
    return jsonify({"ok": True, "count": imported, "order_count": len(created_orders),
                    "imported_rows": imported_rows[:30], "imported_rows_count": len(imported_rows),
                    "skipped": skipped[:30], "skipped_count": len(skipped)})


@app.get("/api/users")
@admin_required
def users_list():
    with db() as conn:
        rows = rows_to_list(conn.execute("SELECT id,username,role,display_name,created_at FROM users ORDER BY id"))
    return jsonify({"ok": True, "items": rows})


@app.post("/api/users")
@admin_required
def users_create():
    data = payload()
    with db() as conn:
        conn.execute(
            "INSERT INTO users (username,password_hash,role,display_name,created_at) VALUES (?,?,?,?,?)",
            (data["username"], generate_password_hash(data["password"]), data.get("role") or "user", data.get("display_name") or data["username"], now_text()),
        )
    log_action("新增用户", data.get("username", ""))
    return jsonify({"ok": True})


execute_schema()


if __name__ == "__main__":
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 8765
    app.run(host="0.0.0.0", port=port, debug=False)
