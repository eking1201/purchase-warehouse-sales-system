from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parents[1]
SOURCE = BASE_DIR / "管理系统.xlsx"
OUTPUT = BASE_DIR / "管理系统_五行示例数据.xlsx"
OUTPUT_DIR = BASE_DIR / "导入示例Excel"


SUPPLIERS = [
    ["SUP0001", "上海精工自动化有限公司", "A", "张强", "13800010001", "上海市浦东新区张江路88号", "工行上海分行 6222****0001", "", "", "私企", "预付30%，发货70%"],
    ["SUP0002", "苏州德瑞机电贸易有限公司", "B", "李娜", "13800010002", "苏州市工业园区星湖街66号", "建行苏州分行 6222****0002", "", "", "私企", "月结"],
    ["SUP0003", "德国贝克传动中国办事处", "A", "王晨", "13800010003", "北京市朝阳区建国路100号", "中行北京分行 6222****0003", "", "", "外企", "预付100%"],
    ["SUP0004", "宁波华兴液压配件厂", "B", "赵磊", "13800010004", "宁波市北仑区富春江路18号", "农行宁波分行 6222****0004", "", "", "私企", "货到付款"],
    ["SUP0005", "昆山宏达五金经营部", "C", "陈敏", "13800010005", "昆山市玉山镇前进路12号", "支付宝 13800010005", "", "", "个人", "现金"],
]

PRODUCTS = [
    ["SP0001", "伺服驱动器", "MR-J4-70A 750W", "个", 8, 2, "", "熔炼", "电气", "12个月", "日本"],
    ["SP0002", "深沟球轴承", "6205 ZZ", "件", 35, 10, "", "加热", "机械", "6个月", "中国"],
    ["SP0003", "液压电磁阀", "DSG-03-3C2-D24", "个", 12, 4, "", "其它", "液体", "12个月", "德国"],
    ["SP0004", "高温密封圈", "FKM 80*95*8", "套", 50, 15, "", "熔炼", "其它", "1个月", "美国"],
    ["SP0005", "温度传感器", "PT100 M8 探头", "个", 20, 6, "", "加热", "电气", "6个月", "中国"],
]

PURCHASE_ORDERS = [
    ["PO202606190001", "上海精工自动化有限公司", "2026-06-19", "SP0001", "伺服驱动器", "MR-J4-70A 750W", "个", 2, 3200, 6400, "", "熔炼", "电气", "12个月", "日本"],
    ["PO202606190002", "苏州德瑞机电贸易有限公司", "2026-06-19", "SP0002", "深沟球轴承", "6205 ZZ", "件", 10, 18, 180, "", "加热", "机械", "6个月", "中国"],
    ["PO202606190003", "德国贝克传动中国办事处", "2026-06-19", "SP0003", "液压电磁阀", "DSG-03-3C2-D24", "个", 3, 860, 2580, "", "其它", "液体", "12个月", "德国"],
    ["PO202606190004", "宁波华兴液压配件厂", "2026-06-19", "SP0004", "高温密封圈", "FKM 80*95*8", "套", 20, 35, 700, "", "熔炼", "其它", "1个月", "美国"],
    ["PO202606190005", "昆山宏达五金经营部", "2026-06-19", "SP0005", "温度传感器", "PT100 M8 探头", "个", 5, 120, 600, "", "加热", "电气", "6个月", "中国"],
]

CUSTOMERS = [
    ["CUS0001", "上海一厂设备部", "A", "刘工", "13900020001", "上海市宝山区铁山路9号", "工行上海分行 6222****1001", "", "", "私企", "月结", 9800, 8000],
    ["CUS0002", "苏州精密制造有限公司", "B", "周经理", "13900020002", "苏州市吴中区金枫路99号", "建行苏州分行 6222****1002", "", "", "私企", "预付30%，发货70%", 4200, 4200],
    ["CUS0003", "杭州热处理中心", "A", "钱工", "13900020003", "杭州市萧山区建设一路18号", "中行杭州分行 6222****1003", "", "", "外企", "预付100%", 7600, 7600],
    ["CUS0004", "宁波压铸有限公司", "B", "孙工", "13900020004", "宁波市镇海区骆驼街道66号", "农行宁波分行 6222****1004", "", "", "私企", "货到付款", 3500, 2000],
    ["CUS0005", "无锡设备维修站", "C", "吴师傅", "13900020005", "无锡市新吴区长江路28号", "微信 13900020005", "", "", "个人", "现金", 1200, 1200],
]

SALES_ORDERS = [
    ["SP0001", "伺服驱动器", "MR-J4-70A 750W", "个", 1, 3900, 3900, "2026-06-25"],
    ["SP0002", "深沟球轴承", "6205 ZZ", "件", 6, 30, 180, "2026-06-22"],
    ["SP0003", "液压电磁阀", "DSG-03-3C2-D24", "个", 2, 1200, 2400, "2026-06-28"],
    ["SP0004", "高温密封圈", "FKM 80*95*8", "套", 12, 58, 696, "2026-06-23"],
    ["SP0005", "温度传感器", "PT100 M8 探头", "个", 4, 180, 720, "2026-06-24"],
]


def copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format


def write_rows(ws, start_row: int, rows: list[list], max_col: int) -> None:
    template_row = start_row
    for offset, row in enumerate(rows):
        target_row = start_row + offset
        copy_row_style(ws, template_row, target_row, max_col)
        for col, value in enumerate(row, start=1):
            ws.cell(target_row, col).value = value


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    section_fill = PatternFill("solid", fgColor="FFF2CC")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.value and "模版" in str(cell.value):
                cell.font = Font(bold=True, size=13)
                cell.fill = section_fill
            elif cell.row in (2, 3, 8, 9, 12, 13):
                if cell.value:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


def add_import_sheet(wb, title: str, headers: list[str], rows: list[list]) -> None:
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def rebuild_workbook() -> None:
    wb = load_workbook(SOURCE)

    ws = wb["采购"]
    write_rows(ws, 4, SUPPLIERS, 11)
    write_rows(ws, 10, PURCHASE_ORDERS, 15)
    ws.cell(15, 9).value = "总计"
    ws.cell(15, 10).value = "=SUM(J10:J14)"
    style_sheet(ws)

    ws = wb["仓库"]
    write_rows(ws, 3, PRODUCTS, 11)
    style_sheet(ws)

    ws = wb["销售"]
    write_rows(ws, 4, CUSTOMERS, 13)
    write_rows(ws, 14, SALES_ORDERS, 8)
    ws.cell(19, 6).value = "总计"
    ws.cell(19, 7).value = "=SUM(G14:G18)"
    style_sheet(ws)

    ws = wb["统计"]
    ws["A3"] = "单个产品采购的统计"
    ws["B3"] = "示例：SP0001 采购数量 2，采购金额 6400"
    ws["A4"] = "单个产品销售统计"
    ws["B4"] = "示例：SP0001 销售数量 1，销售金额 3900"
    ws["A6"] = "单个订单利润统计"
    ws["B6"] = "示例：销售金额 3900 - 估算采购成本 3200 = 利润 700"
    ws["A7"] = "单个客户销售统计"
    ws["B7"] = "示例：上海一厂设备部 应付金额 9800，实收金额 8000"
    style_sheet(ws)

    add_import_sheet(
        wb,
        "导入用_供应商",
        ["编号", "名称", "等级", "联系人", "电话", "地址", "账户信息", "照片", "公司资料", "公司类型", "结算时间"],
        SUPPLIERS,
    )
    add_import_sheet(
        wb,
        "导入用_产品",
        ["备件编号", "产品名称", "技术描述", "单位", "当前库存", "安全库存", "照片", "应用设备", "备件类型", "质保时间", "产地"],
        PRODUCTS,
    )
    add_import_sheet(
        wb,
        "导入用_客户",
        ["编号", "名称", "等级", "联系人", "电话", "地址", "账户信息", "照片", "公司资料", "公司类型", "结算时间", "应付金额", "实收金额"],
        CUSTOMERS,
    )
    add_import_sheet(
        wb,
        "导入用_采购订单",
        ["采购单号", "供应商", "采购日期", "备件编号", "产品名称", "技术描述", "单位", "数量", "单价", "合计", "照片", "应用设备", "备件类型", "质保时间", "产地"],
        PURCHASE_ORDERS,
    )
    add_import_sheet(
        wb,
        "导入用_销售订单",
        ["备件编号", "产品名称", "技术描述", "单位", "数量", "单价", "合计", "货期"],
        SALES_ORDERS,
    )

    wb.save(OUTPUT)


def create_single_workbook(title: str, headers: list[str], rows: list[list], note: str) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws["A1"] = note
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].fill = PatternFill("solid", fgColor="FFF2CC")
    ws["A1"].alignment = Alignment(vertical="center", wrap_text=True)
    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{len(rows)+2}"
    OUTPUT_DIR.mkdir(exist_ok=True)
    wb.save(OUTPUT_DIR / f"{title}.xlsx")


def create_separate_import_files() -> None:
    create_single_workbook(
        "01_供应商导入模板_五行示例",
        ["编号", "名称", "等级", "联系人", "电话", "地址", "账户信息", "照片", "公司资料", "公司类型", "结算时间"],
        SUPPLIERS,
        "对应软件：采购 → 供应商管理。可直接用于供应商批量导入测试。",
    )
    create_single_workbook(
        "02_产品导入模板_五行示例",
        ["备件编号", "产品名称", "技术描述", "单位", "当前库存", "安全库存", "照片", "应用设备", "备件类型", "质保时间", "产地"],
        PRODUCTS,
        "对应软件：仓库 → 产品管理。可直接用于产品批量导入测试。",
    )
    create_single_workbook(
        "03_客户导入模板_五行示例",
        ["编号", "名称", "等级", "联系人", "电话", "地址", "账户信息", "照片", "公司资料", "公司类型", "结算时间", "应付金额", "实收金额"],
        CUSTOMERS,
        "对应软件：销售 → 客户管理。可直接用于客户批量导入测试。",
    )
    create_single_workbook(
        "04_采购订单模板_五行示例",
        ["采购单号", "供应商", "采购日期", "备件编号", "产品名称", "技术描述", "单位", "数量", "单价", "合计", "照片", "应用设备", "备件类型", "质保时间", "产地"],
        PURCHASE_ORDERS,
        "对应软件：采购 → 采购订单。供应商可填写供应商名称；采购单号不填时软件会自动生成。",
    )
    create_single_workbook(
        "05_销售订单模板_五行示例",
        ["备件编号", "产品名称", "技术描述", "单位", "数量", "单价", "合计", "货期"],
        SALES_ORDERS,
        "对应软件：销售 → 销售订单。字段按原Excel销售订单模版；导入时可在软件页面选择客户，不选则归到未指定客户。",
    )


if __name__ == "__main__":
    rebuild_workbook()
    create_separate_import_files()
    print(OUTPUT)
    print(OUTPUT_DIR)
